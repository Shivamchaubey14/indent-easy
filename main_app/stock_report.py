"""
Sale & Stock Report (monthly-cycle stock statement) engine.

Per cycle, per MCC/BMC x Product:

    Closing = Opening + Received + StockTransfer(net) - MPP Sale - Damage - Expire

Data sources:
  * Opening        -> previous cycle's finalized Closing, else a current-inventory snapshot
  * Received (NDS) -> GRN from NDS / other company / party (action=GRN) within the range
  * Received (MCC) -> incoming STN receipts from another MCC/BMC (TRANSFER_IN) within the range
  * StockTransfer  -> outgoing STN transfers (TRANSFER_OUT, negative) within the range
  * MPP Sale       -> aggregated from an uploaded SAP sale export (sum of Order Quantity)
  * Damage / Expire-> filled by the admin (download -> fill -> re-upload)

This module is intentionally self-contained; the only cross-module dependency is the
SAP->ProductMappingGroup resolver in views, imported lazily to avoid an import cycle.
"""

import io
import re
from collections import defaultdict
from decimal import Decimal, InvalidOperation

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from django.db.models import Sum
from django.utils import timezone

from .models import (
    BMCOrMCC,
    Cycle,
    Inventory,
    InventoryHistory,
    MonthlyCycle,
    MPPWithCode,
    Product,
    ProductMapping,
    ProductMappingRelation,
    StockReportProduct,
    StockStatement,
    StockStatementEntry,
    MPPSaleAggregate,
)

REPORT_COLUMNS = [
    "Sr.No", "MCC/BMC Name", "Product Name", "Opp.Balance",
    "Received Product NDS/Other Company", "Received Product MCC/BMC",
    "Stock Transfer Other MCC/BMC", "MPP Sale", "Damage Product",
    "Expire Product", "Closing Stock", "Remark's",
]
# 0-based column indexes within a report row (matches REPORT_COLUMNS order).
COL_PRODUCT = 2
COL_OPENING = 3
COL_RECEIVED = 4        # Received Product NDS/Other Company (GRN)
COL_RECEIVED_MCC = 5    # Received Product MCC/BMC (transfer in)
COL_TRANSFER = 6
COL_MPP_SALE = 7
COL_DAMAGE = 8
COL_EXPIRE = 9
COL_CLOSING = 10


# ---------------------------------------------------------------------------------------
# Mapping helpers (SAP export -> internal master data)
# ---------------------------------------------------------------------------------------
def _strip_leading_zeros(value):
    s = str(value or "").strip()
    return s.lstrip("0") or ("0" if s else "")


def resolve_product(material_code, material_desc, cache):
    """SAP material (code + description) -> internal Product, memoised in `cache`.

    Precedence: service-line exclusion -> curated SAP_SALE_ALIASES -> the
    admin-maintained SAP ProductMapping (by material code, then exact SAP name)
    -> direct product code / exact name / loose-normalised name.

    Deliberately NO first-word fuzzy matching (the old shared resolver filed
    'AI Services' under ghee that way).
    """
    key = (str(material_code or "").strip(), str(material_desc or "").strip().upper())
    if key in cache:
        return cache[key]
    code = key[0]
    desc = str(material_desc or "").strip()
    desc_norm = _norm_product(desc)

    # 0) service lines (AI etc.) are not stock — never fill a product row
    if desc_norm in SAP_NON_STOCK:
        cache[key] = None
        return None

    product = None

    # 1) curated sale-export aliases -> the report's product
    alias_target = _SAP_SALE_ALIASES_NORM.get(desc_norm)
    if alias_target:
        product = Product.objects.filter(name__iexact=alias_target).first()

    # 2) the admin-maintained SAP mapping, by material code then exact SAP name
    if product is None:
        mapping = None
        if code:
            mapping = ProductMapping.objects.filter(
                system="SAP", is_active=True, product_code=code).first()
        if mapping is None and desc:
            mapping = ProductMapping.objects.filter(
                system="SAP", is_active=True, product_name__iexact=desc).first()
        if mapping is not None:
            relation = (ProductMappingRelation.objects
                        .filter(product_mapping=mapping)
                        .select_related("group").first())
            group = relation.group if relation else None
            if group is not None:
                indent = group.mappings.filter(
                    product_mapping__system="INDENT_EASY"
                ).select_related("product_mapping").first()
                candidate_names = []
                if indent and indent.product_mapping:
                    candidate_names.append(indent.product_mapping.product_name)
                candidate_names.append(group.name)
                for name in candidate_names:
                    if not name:
                        continue
                    product = Product.objects.filter(name__iexact=name).first()
                    if product:
                        break

    # 3) Direct fallbacks against Product itself: code, exact name, loose name.
    if product is None and code:
        product = Product.objects.filter(product_code=code).first()
    if product is None and desc:
        product = Product.objects.filter(name__iexact=desc).first()
    if product is None and desc_norm:
        norm_cache = cache.get("_norm")
        if norm_cache is None:
            norm_cache = cache["_norm"] = build_product_norm_cache()
        product = norm_cache.get(desc_norm)

    cache[key] = product
    return product


def resolve_bmc(plant, mcc_name, cache, create=True):
    """SAP (Plant code, Mcc Name) -> BMCOrMCC, memoised in `cache`.

    Matches by plant code, then by name (exact / collapsed, so 'RAMSNEHIGHAT'
    finds 'RAM SANEHI GHAT'). A name match with no stored plant code gets the
    plant backfilled. A location in the sale export that's missing from the
    master entirely is CREATED (name + plant) so its sales fill instead of
    silently dropping — that's how RAIBARELI/MAHARAJGANJ rows were vanishing.
    """
    key = (str(plant or "").strip(), str(mcc_name or "").strip().upper())
    if key in cache:
        return cache[key]
    bmc = None
    if key[0]:
        bmc = BMCOrMCC.objects.filter(plant=key[0]).first()
    if bmc is None and key[1]:
        bmc = (BMCOrMCC.objects.filter(name__iexact=key[1]).first()
               or BMCOrMCC.objects.filter(name__icontains=key[1]).first())
        if bmc is None:
            norm = _norm_bmc(key[1])
            if norm:
                bmc = next((b for b in BMCOrMCC.objects.all()
                            if _norm_bmc(b.name) == norm), None)
        if bmc is not None and key[0] and not bmc.plant:
            bmc.plant = key[0]          # remember the plant code for next time
            bmc.save(update_fields=["plant"])
    if bmc is None and create and key[1]:
        bmc = BMCOrMCC.objects.create(name=key[1], plant=key[0] or None)
    cache[key] = bmc
    return bmc


def resolve_mpp(mpp_code, mpp_name, cache):
    key = (str(mpp_code or "").strip(), str(mpp_name or "").strip().upper())
    if key in cache:
        return cache[key]
    mpp = None
    stripped = _strip_leading_zeros(key[0])
    if stripped:
        mpp = (MPPWithCode.objects.filter(mpp_transaction_code=stripped).first()
               or MPPWithCode.objects.filter(mpp_transaction_code=key[0]).first())
    if mpp is None and key[1]:
        mpp = MPPWithCode.objects.filter(name_with_code__icontains=key[1]).first()
    cache[key] = mpp
    return mpp


def _norm_product(name):
    """Loose product-name key so 'Calcium (01_Ltr)' matches 'Calcium 1 Ltr' etc."""
    s = str(name or "").strip().lower()
    s = re.sub(r"[()_/\-.,]", " ", s)       # punctuation -> space
    s = re.sub(r"\b0+(\d)", r"\1", s)        # 01 -> 1, 05 -> 5
    s = re.sub(r"\s+", " ", s).strip()
    return s


# Curated aliases for the SAP SALE EXPORT: material description -> the REPORT's product.
# The sale export must land on the same products the report rows resolve to; SAP names
# several materials differently (and some name-collide with off-report duplicate products
# like 'Dewormr- IIL' / 'Silage'), which left their MPP Sale invisible on the report.
# Keys are matched loosely (case / punctuation / leading zeros ignored).
SAP_SALE_ALIASES = {
    "Ghee - 1 Kg": "MOTHER DAIRY GHEE 1 LTR",          # report row 'Ghee (01_Ltr)'
    "Ghee - 500 GM": "MOTHER DAIRY GHEE 1/2 L",        # report row 'Ghee (500_ML)'
    "IODOPHORE SOLU 200ML": "Iodophor solution",       # report row 'Iodophor Solution 200_ML'
    "CMT DIP CUP": "Teat dip cup..",                   # report row 'Teat dip cup..'
    "Silage": "Silage Kg",                             # report row 'Silage Kg'
    "Dewormr- IIL": "Dwormer IIL_2200 Mg Tab",         # report row 'Dwormer IIL_2200 Mg Tab'
    "Dewormr- Nilzn Sup": "Dewormer- 100ML",           # report row 'Dewormer Nilzan Suspension ML Nos'
    "Dewormr- 150 mg": "Dwormer IIL Calf 150_Mg Tab",  # report row 'Dwormer IIL Calf 150_Mg Tab'
}

# Sale-export lines that are SERVICES, not stock — they must never fill a product row.
# ('AI Services' used to fuzzy-match 'MOTHER D*AI*RY GHEE 1 LTR' and inflate ghee sales.)
SAP_NON_STOCK = {"ai services"}

_SAP_SALE_ALIASES_NORM = {_norm_product(k): v for k, v in SAP_SALE_ALIASES.items()}


# Curated aliases: Sale/Stock-report (Excel) product name  ->  the real product in the
# indent product table. Left-hand keys are matched loosely (case / punctuation / leading
# zeros ignored), so an uploaded "Cattle Feed Sag (50_Kg)" resolves to "CF- SAG 50 KG"
# instead of creating a duplicate product.
PRODUCT_ALIASES = {
    "Cattle Feed Sag (50_Kg)": "CF- SAG 50 KG",
    "Cattle Feed Sag (25_Kg)": "CF- SAG 25 KG",
    "Cattle Feed Gold (50_Kg)": "CF- Gold 50 KG",
    "Cattle Feed Gold (25_Kg)": "CF- Gold 25 KG",
    "Cattle Feed Sag Type_01 (40_Kg)": "CF- SAG 40 KG TYPE1",
    "Cattle Feed Buff.. Spl (40_Kg)": "CF- Buff Spl 40 KG",
    "Pregancy Feed (40_Kg)": "CF-Pregnancy 40 Kg",
    "Early Lactation Feed (40_Kg)": "EARLY LACTATION FEED",
    "Jica Miniral Mixture (01_Kg)": "MM- 1KG JICA",
    "Ghee (01_Ltr)": "MOTHER DAIRY GHEE 1 LTR",
    "Ghee (500_ML)": "MOTHER DAIRY GHEE 1/2 L",
    "Calcium (05_Ltr)": "Calcium 5 Ltr",
    "Calcium (01_Ltr)": "Calcium 1 Ltr",
    "Chelated Mineral Mixture (01_Kg)": "MM- Chealted",
    "Mineral Mixture  (01_Kg)": "MM- 1 KG",
    "Goudhara Shakti (01_Kg)": "Gaudhra Shakti 1Kg",
    "Calsagar Plus (01_Kg)": "Calsagar Plus 1 Kg",
    "Iodophor Solution 200_ML": "Iodophor solution",
    "TSC Nos": "Trisodium citrate",
    "Gulab Jamun (01_kg)": "GULAB JAMUN (1KG)",
    "Dewormer Bolus Tab.": "Dewormer- Bolus",
    "Dewormer Nilzan Suspension ML Nos": "Dewormer- 100ML",
    "Jwar Seed (01_Kg)": "Seeds- Jowar 1 KG",
    "Bazara Seed (01_Kg)": "Seeds- Bazra 1 KG",
    "Sorgam CH43 (03_Kg)": "Seeds- Sorgum 3 KG",
    "Hybrid Jumbo Gold (01_Kg)": "Hyb Jumbo Gold 1Kg",
    "Maze (01_Kg)": "Seeds- Maize 1 KG",
    "Teat Dip Nos": "Teat dip solution",
    # 'Oats Seeds jau (25_Kg)' is the real product; these two name variants fold into it.
    "Oats Seeds (25_Kg)": "Oats Seeds jau (25_Kg)",
    "Jau ( 25 Kg Bag)": "Oats Seeds jau (25_Kg)",
}

# Uploaded rows whose "product" is really a header/blank leaking through — never create these.
_PRODUCT_NAME_JUNK = {"product name", "sr.no", "total", "mcc/bmc name"}

# Reverse of PRODUCT_ALIASES: real indent product name -> the report/SMPCL name to DISPLAY.
# So a stored 'CF- SAG 50 KG' shows as 'Cattle Feed Sag (50_Kg)' in the report and workbook.
# Skip the name-variant aliases whose target is itself a report-style product (oats).
REVERSE_ALIAS = {}
for _ex, _real in PRODUCT_ALIASES.items():
    if _ex.strip().lower() in ("oats seeds (25_kg)", "jau ( 25 kg bag)"):
        continue
    REVERSE_ALIAS.setdefault(_real.strip().lower(), _ex)


def report_display_name(product_name):
    """The name to show in the report — the report/SMPCL alias if the product has one,
    otherwise the product's own name."""
    return REVERSE_ALIAS.get(str(product_name or "").strip().lower(), product_name)


# The fixed Sale & Stock Report product list, in the standard order the report uses.
# Names are the report/SMPCL names; each resolves (via alias / exact) to a real product.
REPORT_ORDER = [
    "Cattle Feed Sag (50_Kg)", "Cattle Feed Sag (25_Kg)", "Cattle Feed Gold (50_Kg)",
    "Cattle Feed Gold (25_Kg)", "Cattle Feed Sag Type_01 (40_Kg)", "Cattle Feed Buff.. Spl (40_Kg)",
    "Pregancy Feed (40_Kg)", "Early Lactation Feed (40_Kg)", "Jica Miniral Mixture (01_Kg)",
    "Ghee (01_Ltr)", "Ghee (500_ML)", "Calcium (05_Ltr)", "Calcium (01_Ltr)",
    "Chelated Mineral Mixture (01_Kg)", "Mineral Mixture  (01_Kg)", "Goudhara Shakti (01_Kg)",
    "Calsagar Plus (01_Kg)", "Dewormer Bolus Tab.", "Dewormer Nilzan Suspension ML Nos",
    "Dwormer IIL_2200 Mg Tab", "Dwormer IIL Calf 150_Mg Tab", "Gulab Jamun (01_kg)",
    "Rasgulla (01_Kg)", "Iodophor Solution 200_ML", "Teat Dip Nos", "TSC Nos",
    "Mustard 250g + Berseem 02_Kg", "Oats Seeds jau (25_Kg)", "Jwar Seed (01_Kg)",
    "Bazara Seed (01_Kg)", "Sorgam CH43 (03_Kg)", "Hybrid Jumbo Gold (01_Kg)", "Maze (01_Kg)",
    "Silage Kg", "Teat dip cup..",
]


def report_products():
    """Ordered [(Product, display_name)] for the report — the fixed REPORT_ORDER set plus
    any admin-added StockReportProduct extras, each inserted at its saved position."""
    cache = build_product_norm_cache()
    out, seen = [], set()
    for smpcl in REPORT_ORDER:
        p = resolve_product_by_name(smpcl, cache, create=False)
        if p and p.id not in seen:
            out.append((p, smpcl))
            seen.add(p.id)
    for extra in StockReportProduct.objects.select_related("product").order_by("position", "id"):
        p = extra.product
        if p.id in seen:
            continue
        pos = min(max(extra.position, 0), len(out))
        out.insert(pos, (p, extra.display_name or p.name))
        seen.add(p.id)
    return out


def report_product_index():
    """{real product-name (lower): (order, display_name)} for the report product list.
    Used to restrict and order the report to the standard product set."""
    idx = {}
    for pos, (p, display) in enumerate(report_products()):
        idx.setdefault(p.name.strip().lower(), (pos, display))
    return idx


def report_location_restrictions():
    """{product_id: set(bmc_ids)} for admin-added report products that are limited to
    specific locations. Products absent from the dict show for every location."""
    restrict = {}
    for extra in StockReportProduct.objects.prefetch_related("locations"):
        loc_ids = {b.id for b in extra.locations.all()}
        if loc_ids:
            restrict[extra.product_id] = loc_ids
    return restrict


def ensure_all_locations(statement):
    """Create zero entries so EVERY MCC/BMC shows EVERY report product, even a location with
    no data (so e.g. DURJANPURGHAT still appears with all zeros). Additive only — never
    touches existing entries, so it's safe on a manually-overridden statement.
    Admin-added products limited to specific locations are only created there."""
    pids = [p.id for p, _d in report_products()]
    if not pids:
        return 0
    restrict = report_location_restrictions()
    existing = {(e.bmc_or_mcc_id, e.product_id) for e in statement.entries.all()}
    to_create = [
        StockStatementEntry(statement=statement, bmc_or_mcc_id=b.id, product_id=pid)
        for b in BMCOrMCC.objects.all()
        for pid in pids
        if (b.id, pid) not in existing
        and (pid not in restrict or b.id in restrict[pid])
    ]
    if to_create:
        StockStatementEntry.objects.bulk_create(to_create, batch_size=1000)
    return len(to_create)


def _norm_bmc(name):
    """Collapse a BMC/MCC name for matching so 'RAMSANEHI GHAT' == 'RAM SANEHI GHAT'."""
    return re.sub(r"[^a-z0-9]", "", str(name or "").lower())


def _map_report_columns(header):
    """Map report columns by HEADER TEXT so uploads work whatever the layout:
    single 'Received' or split 'Received NDS/Other Company' + 'Received MCC/BMC', and
    combined 'Damage Product / Expire Item' or separate 'Damage'/'Expire'.
    `header` is the header row (list of cell values). Returns {field: 1-based col index}."""
    cols = {}
    for i, raw in enumerate(header, start=1):
        h = re.sub(r"\s+", " ", str(raw or "").strip().lower())
        if not h:
            continue
        if "product name" in h:
            cols.setdefault("product", i)
        elif ("mcc" in h or "bmc" in h) and "name" in h:
            cols.setdefault("bmc", i)
        elif "opp" in h or "opening" in h:
            cols.setdefault("opening", i)
        elif "received" in h and "mcc" in h:
            cols.setdefault("received_mcc", i)
        elif "received" in h:                       # NDS/Other Company, or a single 'Received'
            cols.setdefault("received", i)
        elif "transfer" in h:
            cols.setdefault("transfer", i)
        elif "sale" in h:
            cols.setdefault("sale", i)
        elif "damage" in h and "expire" in h:
            cols.setdefault("damage_combined", i)
        elif "damage" in h:
            cols.setdefault("damage", i)
        elif "expire" in h:
            cols.setdefault("expire", i)
        elif "closing" in h:
            cols.setdefault("closing", i)
        elif "remark" in h:
            cols.setdefault("remark", i)
    return cols


def build_product_norm_cache():
    """Map loose-normalised product name -> Product for pattern matching, then overlay the
    curated PRODUCT_ALIASES so report names resolve to the real indent products."""
    cache = {}
    by_lower = {}
    for p in Product.objects.all().only("id", "name"):
        cache.setdefault(_norm_product(p.name), p)
        by_lower[p.name.strip().lower()] = p
    for excel_name, real_name in PRODUCT_ALIASES.items():
        target = by_lower.get(real_name.strip().lower()) or cache.get(_norm_product(real_name))
        if target is not None:
            cache[_norm_product(excel_name)] = target   # alias wins over any loose match
    return cache


def resolve_product_by_name(name, norm_cache, create=False):
    """Match an uploaded product name to a Product: curated alias / exact (case-insensitive) /
    loose pattern. If ``create`` and nothing matches, create a Product with just the name
    (all other fields left blank / None). Header/junk rows never match or create."""
    raw = str(name or "").strip()
    if not raw or raw.lower() in _PRODUCT_NAME_JUNK:
        return None
    key = _norm_product(raw)
    if key and key in norm_cache:      # alias or previously-cached match
        return norm_cache[key]
    p = Product.objects.filter(name__iexact=raw).first()
    if p:
        return p
    if create:
        p = Product.objects.create(name=raw)   # unique name; other fields blank/default/None
        if key:
            norm_cache[key] = p
        return p
    return None


# ---------------------------------------------------------------------------------------
# Cycle helpers
# ---------------------------------------------------------------------------------------
_MONTHS = ["january", "february", "march", "april", "may", "june", "july",
           "august", "september", "october", "november", "december"]

_MONTH_ABBR = {m[:3]: m.capitalize() for m in _MONTHS}


def parse_cycle_header(text):
    """Read a sheet's cycle line, e.g. 'Sale Report Cycle-01-10 Apr-2026' (or
    '... Sale Report Cycle 01-10 Jul 2026'), returning (start_day, end_day, 'April', 2026)
    or None if the text isn't a cycle header."""
    m = re.search(
        r"sale\s*report\s*cycle[\s:_\-]*?(\d{1,2})\s*[-–]\s*(\d{1,2})\s+([A-Za-z]{3,})[\s\-_]*?(\d{4})",
        str(text or ""), re.I)
    if not m:
        return None
    sd, ed = int(m.group(1)), int(m.group(2))
    mon = m.group(3).strip().lower()[:3]
    return sd, ed, _MONTH_ABBR.get(mon, m.group(3).capitalize()), int(m.group(4))


def detect_cycle_from_workbook(wb):
    """Identify which Cycle a report workbook belongs to from its 'Sale Report Cycle ...'
    header. Returns (Cycle or None, parsed_tuple or None). parsed is set even when no
    matching Cycle exists, so the caller can report a helpful message."""
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
            for v in row:
                parsed = parse_cycle_header(v)
                if parsed:
                    sd, ed, month, yr = parsed
                    cyc = (Cycle.objects
                           .filter(monthly_cycle__month__iexact=month, monthly_cycle__year=yr,
                                   start_date__day=sd, end_date__day=ed)
                           .select_related("monthly_cycle").first())
                    return cyc, parsed
    return None, None


def get_previous_cycle(cycle):
    """The cycle immediately before `cycle` (same month; else last cycle of prev month)."""
    mc = cycle.monthly_cycle
    prev = Cycle.objects.filter(
        monthly_cycle=mc, cycle_number=cycle.cycle_number - 1
    ).first()
    if prev:
        return prev
    if cycle.cycle_number <= 1 and mc is not None:
        month_name = (mc.month or "").strip().lower()
        try:
            idx = _MONTHS.index(month_name)
        except ValueError:
            return None
        if idx == 0:
            prev_month, prev_year = "December", (mc.year or 0) - 1
        else:
            prev_month, prev_year = _MONTHS[idx - 1].capitalize(), mc.year
        return (Cycle.objects.filter(
            monthly_cycle__month__iexact=prev_month,
            monthly_cycle__year=prev_year,
        ).order_by("-cycle_number").first())
    return None


# ---------------------------------------------------------------------------------------
# Step 1: generate the statement (Opening / Received / Stock Transfer)
# ---------------------------------------------------------------------------------------
def generate_statement(cycle, user=None):
    """Create/refresh a StockStatement for `cycle`. Preserves any already-entered
    MPP Sale / Damage / Expire and just refreshes Opening / Received / Stock Transfer."""
    statement, _ = StockStatement.objects.get_or_create(cycle=cycle)

    # location string -> BMCOrMCC (Inventory is keyed by user.location, which equals BMCOrMCC.name)
    bmc_by_name = {b.name.strip().upper(): b for b in BMCOrMCC.objects.all()}

    def bmc_for_location(loc):
        return bmc_by_name.get(str(loc or "").strip().upper())

    # ---- Opening: carry the previous cycle's closing forward ----
    # Within a month, always chain to the previous cycle. ACROSS months, only carry from a
    # CONFIRMED month (uploaded / finalized) — so an uploaded month seeds the next month's
    # first cycle, but empty future months stay at 0 instead of inheriting stale data or an
    # unrelated inventory snapshot (which was bleeding April's numbers into June/July).
    opening = defaultdict(int)   # (bmc_id, product_id) -> qty
    prev = get_previous_cycle(cycle)
    prev_stmt = getattr(prev, "stock_statement", None) if prev else None
    same_month = bool(prev and prev.monthly_cycle_id == cycle.monthly_cycle_id)
    confirmed = prev_stmt is not None and (
        prev_stmt.is_manual_override or prev_stmt.status == StockStatement.STATUS_FINALIZED)
    if prev_stmt is not None and (same_month or confirmed):
        for e in prev_stmt.entries.all():
            opening[(e.bmc_or_mcc_id, e.product_id)] = e.closing_balance
    elif prev is None:
        # genuinely the first cycle in the system -> seed from current inventory
        snap = (Inventory.objects
                .values("mcc_bmc_user__location", "product_id")
                .annotate(q=Sum("quantity")))
        for row in snap:
            bmc = bmc_for_location(row["mcc_bmc_user__location"])
            if bmc and row["product_id"] and row["q"]:
                opening[(bmc.id, row["product_id"])] += int(row["q"] or 0)
    # else: a cross-month link to an unconfirmed prior month -> opening stays 0 (empty)

    # ---- Received (split) and Stock Transfer from InventoryHistory in range ----
    # received      = GRN from NDS / other company / party (action=GRN)
    # received_mcc  = incoming STN transfer received from another MCC/BMC (TRANSFER_IN, +ve)
    # transfer      = only OUTGOING transfers to other MCC/BMC (TRANSFER_OUT, negative)
    received = defaultdict(int)
    received_mcc = defaultdict(int)
    transfer = defaultdict(int)
    hist = (InventoryHistory.objects
            .filter(created_at__date__range=[cycle.start_date, cycle.end_date])
            .values("inventory__mcc_bmc_user__location", "inventory__product_id", "action")
            .annotate(q=Sum("quantity_change")))
    for row in hist:
        bmc = bmc_for_location(row["inventory__mcc_bmc_user__location"])
        pid = row["inventory__product_id"]
        if not bmc or not pid:
            continue
        k = (bmc.id, pid)
        action = row["action"]
        qty = int(row["q"] or 0)
        if action == "GRN":
            received[k] += qty
        elif action == "TRANSFER_IN":
            received_mcc[k] += qty
        elif action == "TRANSFER_OUT":
            transfer[k] += qty   # -ve; only stock sent out to other MCC/BMC

    # ---- Upsert entries: FULL product grid (preserve sale/damage/expire) ----
    combos = set(opening) | set(received) | set(received_mcc) | set(transfer)
    existing = {(e.bmc_or_mcc_id, e.product_id): e for e in statement.entries.all()}
    combos |= set(existing.keys())
    # Every location that appears in the cycle shows EVERY product that appears in the cycle
    # (zero rows included), so each sheet matches the standard master layout.
    present_bmcs = {b for (b, _p) in combos}
    present_products = {p for (_b, p) in combos}
    if present_bmcs and present_products:
        combos = {(b, p) for b in present_bmcs for p in present_products}

    to_create, to_update = [], []
    for (bmc_id, pid) in combos:
        o = opening.get((bmc_id, pid), 0)
        r = received.get((bmc_id, pid), 0)
        rm = received_mcc.get((bmc_id, pid), 0)
        t = transfer.get((bmc_id, pid), 0)
        entry = existing.get((bmc_id, pid))
        if entry is None:
            entry = StockStatementEntry(statement=statement, bmc_or_mcc_id=bmc_id, product_id=pid,
                                        opening_balance=o, received=r, received_mcc=rm, stock_transfer=t)
            entry.closing_balance = entry.compute_closing()
            to_create.append(entry)
        else:
            entry.opening_balance = o; entry.received = r
            entry.received_mcc = rm; entry.stock_transfer = t
            entry.closing_balance = entry.compute_closing()   # preserves sale/damage/expire
            to_update.append(entry)
    if to_create:
        StockStatementEntry.objects.bulk_create(to_create, batch_size=500)
    if to_update:
        StockStatementEntry.objects.bulk_update(
            to_update, ["opening_balance", "received", "received_mcc", "stock_transfer", "closing_balance"],
            batch_size=500)

    if statement.status == StockStatement.STATUS_FINALIZED:
        statement.status = StockStatement.STATUS_SALE_UPLOADED
    statement.generated_by = user or statement.generated_by
    statement.generated_at = timezone.now()
    statement.is_manual_override = False   # recomputed from source -> drops any manual override
    statement.save()
    ensure_all_locations(statement)   # every location shows every report product (even zeros)
    return statement


def rechain_month(monthly_cycle):
    """Re-link every cycle of a month so each cycle's Opening = the previous cycle's Closing
    (per BMC/MCC + product), recomputing closings forward. The first cycle's Opening is left
    as-is (inventory snapshot / manual). Called whenever a cycle's numbers change so later
    cycles auto-fill from the corrected/updated closing.
    """
    if monthly_cycle is None:
        return
    stmts = list(StockStatement.objects
                 .filter(cycle__monthly_cycle=monthly_cycle)
                 .select_related("cycle")
                 .order_by("cycle__cycle_number"))
    for i in range(1, len(stmts)):
        prev, cur = stmts[i - 1], stmts[i]
        closings = {(e.bmc_or_mcc_id, e.product_id): e.closing_balance for e in prev.entries.all()}
        to_update = []
        for e in cur.entries.all():
            new_open = closings.get((e.bmc_or_mcc_id, e.product_id), 0)
            if e.opening_balance != new_open:
                e.opening_balance = new_open
                e.closing_balance = e.compute_closing()
                to_update.append(e)
        if to_update:
            StockStatementEntry.objects.bulk_update(
                to_update, ["opening_balance", "closing_balance"], batch_size=500)


# ---------------------------------------------------------------------------------------
# Step 2: ingest the SAP sale export -> MPP Sale
# ---------------------------------------------------------------------------------------
def _norm_header(v):
    return re.sub(r"\s+", " ", str(v or "").strip().lower())


def _materialize(ws):
    """Read a worksheet's rows once as a list of tuples.

    IMPORTANT: openpyxl read-only worksheets stream from the top, so random
    ``ws.cell(r, c)`` access is O(n) per call -> O(n**2) over a loop and hangs on
    files with a few thousand rows. Reading everything once with ``iter_rows`` keeps
    ingestion O(n). Access a cell as ``row[idx - 1]`` (1-based column index).
    """
    return [tuple(r) for r in ws.iter_rows(values_only=True)]


def _cell(row, idx):
    """1-based column access into a materialized row tuple (None if out of range)."""
    if not idx:
        return None
    i = idx - 1
    return row[i] if 0 <= i < len(row) else None


def _find_columns(rows):
    """Locate the header row (containing 'Plant') and map our fields to column indexes.

    `rows` is a list of row tuples (from :func:`_materialize`). Returns the 0-based
    header-row index into `rows` and the field->column-index (1-based) map.
    """
    header_row = None
    headers = None
    for idx, row in enumerate(rows[:15]):
        values = [_norm_header(v) for v in row]
        if "plant" in values:
            header_row = idx
            headers = values
            break
    if header_row is None:
        raise ValueError("Could not find a header row containing 'Plant' in the sale file.")

    def col(*needles):
        for idx, h in enumerate(headers, start=1):
            if all(n in h for n in needles):
                return idx
        return None

    cols = {
        "plant": col("plant") if "plant" in headers else col("plant"),
        "mcc": col("mcc", "name"),
        "mpp_name": col("mpp", "name"),
        "mpp_code": col("mpp", "code"),
        "date": col("document", "date"),
        "material_code": col("material") if col("material") and not col("material", "desc") else col("material"),
        "material_desc": col("material", "desc") or col("description"),
        "qty": col("order", "quantity") or col("quantity"),
        "net": col("net", "value"),
    }
    # 'Material' (code) vs 'Material Description': pick the plain 'material' column for code.
    for idx, h in enumerate(headers, start=1):
        if h == "material":
            cols["material_code"] = idx
    return header_row, cols


def ingest_sale(statement, file_obj, filename=""):
    """Parse an uploaded SAP sale export and fill MPP Sale per (BMC/MCC, Product)."""
    wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
    ws = wb.active
    rows = _materialize(ws)
    wb.close()
    header_row, cols = _find_columns(rows)

    prod_cache, bmc_cache, mpp_cache = {}, {}, {}
    # aggregation keys
    per_entry = defaultdict(int)                 # (bmc_id, product_id) -> qty
    per_mpp = {}                                 # (mpp_code, material_code) -> dict
    total = matched = unmatched = services = 0

    for r in range(header_row + 1, len(rows)):
        row = rows[r]
        plant = _cell(row, cols["plant"])
        mcc_name = _cell(row, cols["mcc"])
        mpp_code = _cell(row, cols["mpp_code"])
        mpp_name = _cell(row, cols["mpp_name"])
        material_code = _cell(row, cols["material_code"])
        material_desc = _cell(row, cols["material_desc"])
        qty_raw = _cell(row, cols["qty"])
        net_raw = _cell(row, cols["net"])
        if not any(str(v or "").strip() for v in (plant, mcc_name, material_code, material_desc)):
            continue   # blank row, or the trailing grand-total row (qty with no identifiers)
        if _norm_product(material_desc) in SAP_NON_STOCK:
            services += 1   # service lines (AI etc.) aren't products — leave them out entirely
            continue
        total += 1

        try:
            qty = int(round(float(qty_raw))) if qty_raw not in (None, "") else 0
        except (TypeError, ValueError):
            qty = 0
        try:
            net_value = Decimal(str(net_raw)) if net_raw not in (None, "") else Decimal("0")
        except (InvalidOperation, TypeError):
            net_value = Decimal("0")

        product = resolve_product(material_code, material_desc, prod_cache)
        bmc = resolve_bmc(plant, mcc_name, bmc_cache)
        mpp = resolve_mpp(mpp_code, mpp_name, mpp_cache)
        is_matched = bool(product and bmc)
        if is_matched:
            matched += 1
            per_entry[(bmc.id, product.id)] += qty
        else:
            unmatched += 1

        agg_key = (str(mpp_code or "").strip(), str(material_code or "").strip())
        row = per_mpp.get(agg_key)
        if row is None:
            per_mpp[agg_key] = {
                "bmc": bmc, "mpp": mpp, "product": product,
                "sap_plant": str(plant or "").strip(), "sap_mcc_name": str(mcc_name or "").strip(),
                "sap_mpp_code": str(mpp_code or "").strip(), "sap_mpp_name": str(mpp_name or "").strip(),
                "sap_material_code": str(material_code or "").strip(),
                "sap_material_desc": str(material_desc or "").strip(),
                "qty": qty, "net_value": net_value, "matched": is_matched,
            }
        else:
            row["qty"] += qty
            row["net_value"] += net_value

    # ---- persist MPP-level aggregation (rebuild) ----
    statement.mpp_sales.all().delete()
    bulk = []
    for row in per_mpp.values():
        bulk.append(MPPSaleAggregate(
            statement=statement,
            bmc_or_mcc=row["bmc"], mpp=row["mpp"], product=row["product"],
            sap_plant=row["sap_plant"][:50], sap_mcc_name=row["sap_mcc_name"][:150],
            sap_mpp_code=row["sap_mpp_code"][:50], sap_mpp_name=row["sap_mpp_name"][:150],
            sap_material_code=row["sap_material_code"][:50],
            sap_material_desc=row["sap_material_desc"][:255],
            quantity=row["qty"], net_value=row["net_value"], matched=row["matched"],
        ))
    MPPSaleAggregate.objects.bulk_create(bulk, batch_size=500)

    # ---- apply MPP Sale onto entries (reset then set; create combos as needed) ----
    statement.entries.exclude(mpp_sale=0).update(mpp_sale=0)   # clear stale sale values
    for (bmc_id, pid), qty in per_entry.items():
        entry, _ = StockStatementEntry.objects.get_or_create(
            statement=statement, bmc_or_mcc_id=bmc_id, product_id=pid,
        )
        entry.mpp_sale = qty
        entry.save()
    # recompute closing for any entry whose sale was reset to 0
    for entry in statement.entries.all():
        new_closing = entry.compute_closing()
        if entry.closing_balance != new_closing:
            entry.closing_balance = new_closing
            entry.save(update_fields=["closing_balance", "updated_at"])

    statement.sale_file_name = filename or getattr(file_obj, "name", "")
    statement.sale_rows_total = total
    statement.sale_rows_matched = matched
    statement.sale_rows_unmatched = unmatched
    statement.sale_uploaded_at = timezone.now()
    if statement.status == StockStatement.STATUS_DRAFT:
        statement.status = StockStatement.STATUS_SALE_UPLOADED
    statement.save()
    # Locations first seen in this sale file (auto-created above) get their full
    # zero-filled product grid so their sheet shows every report product.
    ensure_all_locations(statement)

    return {
        "total": total, "matched": matched, "unmatched": unmatched,
        "services_skipped": services,
        "mpp_rows": len(per_mpp), "entries_with_sale": len(per_entry),
    }


# ---------------------------------------------------------------------------------------
# Step 3: build the downloadable report (one sheet per MCC/BMC, in the required format)
# ---------------------------------------------------------------------------------------
_THIN = Side(style="thin", color="B0B0B0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")


def _safe_sheet_title(name, used):
    title = re.sub(r"[\\/*?:\[\]]", "-", str(name or "Sheet")).strip()[:31] or "Sheet"
    base, i = title, 1
    while title in used:
        suffix = f"_{i}"
        title = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(title)
    return title


def build_workbook(statement, only_bmc=None, include_summary=False):
    """Build the formatted report workbook (one sheet per MCC/BMC).

    Pass ``only_bmc`` (a BMCOrMCC instance) to restrict the workbook to a single
    location's sheet — used by the location-level download.
    Pass ``include_summary=True`` (admin full download) to append the company-wide
    'SMPCL SUMMERY' sheet: a month-wise rollup across all locations (opening from the
    first cycle, closing from the last cycle, flows summed across the month's cycles).
    """
    cycle = statement.cycle
    month_display = ""
    if cycle.monthly_cycle:
        month_display = f"{cycle.monthly_cycle.month}-{cycle.monthly_cycle.year}"
    cycle_range = ""
    if cycle.start_date and cycle.end_date:
        cycle_range = f"{cycle.start_date.strftime('%d')}-{cycle.end_date.strftime('%d %b %Y')}"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used_titles = set()

    entries = (statement.entries
               .select_related("bmc_or_mcc", "product")
               .order_by("bmc_or_mcc__name", "product__name"))
    if only_bmc is not None:
        entries = entries.filter(bmc_or_mcc=only_bmc)
    report_idx = report_product_index()   # restrict + order to the standard report products
    by_bmc = defaultdict(list)
    for e in entries:
        ri = report_idx.get(e.product.name.strip().lower())
        if ri is None:
            continue
        by_bmc[e.bmc_or_mcc].append((ri[0], e))
    for bmc in by_bmc:
        by_bmc[bmc].sort(key=lambda oe: oe[0])
        by_bmc[bmc] = [e for _o, e in by_bmc[bmc]]

    if not by_bmc:
        ws = wb.create_sheet(_safe_sheet_title("No Data", used_titles))
        ws["A1"] = "No stock statement entries. Generate the statement first."
        bio = io.BytesIO(); wb.save(bio); bio.seek(0)
        return bio

    for bmc in sorted(by_bmc.keys(), key=lambda b: b.name):
        ws = wb.create_sheet(_safe_sheet_title(bmc.name, used_titles))
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(REPORT_COLUMNS))
        ws.cell(1, 1, "Shwetdhara Milk Producer Company Limited").font = Font(bold=True, size=13)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(REPORT_COLUMNS))
        ws.cell(2, 1, f"Input Sale & Stock Report  Month {month_display}").font = Font(bold=True)
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(REPORT_COLUMNS))
        ws.cell(3, 1, f"{bmc.name}  |  Sale Report Cycle {cycle_range}").font = Font(bold=True)

        # header row (row 4)
        for c, title in enumerate(REPORT_COLUMNS, start=1):
            cell = ws.cell(4, c, title)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _BORDER

        totals = defaultdict(int)
        row_ptr = 5
        for i, e in enumerate(by_bmc[bmc], start=1):
            values = [
                i, bmc.name, report_display_name(e.product.name), e.opening_balance, e.received, e.received_mcc,
                e.stock_transfer, e.mpp_sale, e.damage, e.expire, e.closing_balance, e.remark,
            ]
            for c, v in enumerate(values, start=1):
                cell = ws.cell(row_ptr, c, v)
                cell.border = _BORDER
                if c >= 4:
                    cell.alignment = Alignment(horizontal="center")
            for col in (COL_OPENING, COL_RECEIVED, COL_RECEIVED_MCC, COL_TRANSFER, COL_MPP_SALE,
                        COL_DAMAGE, COL_EXPIRE, COL_CLOSING):
                totals[col] += values[col] or 0
            row_ptr += 1

        # total row
        tcell = ws.cell(row_ptr, 1, "Total")
        tcell.font = Font(bold=True)
        for c in range(1, len(REPORT_COLUMNS) + 1):
            cell = ws.cell(row_ptr, c)
            cell.border = _BORDER
            cell.fill = _TOTAL_FILL
            cell.font = Font(bold=True)
            col0 = c - 1
            if col0 in totals:
                cell.value = totals[col0]
                cell.alignment = Alignment(horizontal="center")

        widths = [7, 22, 34, 12, 18, 16, 16, 11, 13, 13, 13, 16]
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A5"

    if include_summary and only_bmc is None:
        build_smpcl_summary(wb, statement, used_titles)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def build_smpcl_summary(wb, statement, used_titles):
    """Append the company-wide 'SMPCL SUMMERY' sheet — one row per product, aggregated
    across ALL locations for the whole month:

        * Opp.Balance -> opening of the FIRST cycle of the month (summed over locations)
        * Closing     -> closing of the LAST cycle of the month (summed over locations)
        * Received / Transfer / MPP Sale / Damage / Expire -> summed across every cycle
    """
    cycle = statement.cycle
    mc = cycle.monthly_cycle
    stmts = list(StockStatement.objects
                 .filter(cycle__monthly_cycle=mc)
                 .select_related("cycle")
                 .order_by("cycle__cycle_number")) if mc else [statement]
    if not stmts:
        stmts = [statement]
    first_stmt, last_stmt = stmts[0], stmts[-1]
    stmt_ids = [s.id for s in stmts]

    agg = {}   # product_id -> aggregated dict

    def _row(pid):
        return agg.setdefault(pid, {"opening": 0, "recv": 0, "recv_mcc": 0, "xfer": 0,
                                    "sale": 0, "damage": 0, "expire": 0, "closing": 0})

    # flows summed across every cycle of the month
    for r in (StockStatementEntry.objects.filter(statement_id__in=stmt_ids)
              .values("product_id")
              .annotate(recv=Sum("received"), recv_mcc=Sum("received_mcc"),
                        xfer=Sum("stock_transfer"), sale=Sum("mpp_sale"),
                        damage=Sum("damage"), expire=Sum("expire"))):
        d = _row(r["product_id"])
        d["recv"] = r["recv"] or 0; d["recv_mcc"] = r["recv_mcc"] or 0; d["xfer"] = r["xfer"] or 0
        d["sale"] = r["sale"] or 0; d["damage"] = r["damage"] or 0; d["expire"] = r["expire"] or 0
    # opening from the first cycle, closing from the last cycle
    for r in (StockStatementEntry.objects.filter(statement=first_stmt)
              .values("product_id").annotate(o=Sum("opening_balance"))):
        _row(r["product_id"])["opening"] = r["o"] or 0
    for r in (StockStatementEntry.objects.filter(statement=last_stmt)
              .values("product_id").annotate(c=Sum("closing_balance"))):
        _row(r["product_id"])["closing"] = r["c"] or 0

    products = {p.id: p for p in Product.objects.filter(id__in=agg.keys())}
    report_idx = report_product_index()   # restrict + order to the standard report products
    _tmp = []
    for pid, d in agg.items():
        p = products.get(pid)
        ri = report_idx.get(p.name.strip().lower()) if p else None
        if ri is not None:
            _tmp.append((ri[0], pid, d))
    _tmp.sort(key=lambda x: x[0])
    ordered = [(pid, d) for _o, pid, d in _tmp]

    ncol = len(REPORT_COLUMNS)
    month_display = f"{mc.month}-{mc.year}" if mc else ""
    ws = wb.create_sheet(_safe_sheet_title("SMPCL SUMMERY", used_titles))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    ws.cell(1, 1, "Shwetdhara Milk Producer Company Limited").font = Font(bold=True, size=13)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    ws.cell(2, 1, f"Input Sale & Stock Report Summery  Month {month_display}").font = Font(bold=True)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncol)
    ws.cell(3, 1, "SMPCL STORE  |  All MCC/BMC locations").font = Font(bold=True)

    for c, title in enumerate(REPORT_COLUMNS, start=1):
        cell = ws.cell(4, c, title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER

    totals = defaultdict(int)
    row_ptr = 5
    for i, (pid, d) in enumerate(ordered, start=1):
        p = products.get(pid)
        values = [
            i, "SMPCL STORE", (report_display_name(p.name) if p else str(pid)), d["opening"], d["recv"], d["recv_mcc"],
            d["xfer"], d["sale"], d["damage"], d["expire"], d["closing"], "",
        ]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row_ptr, c, v)
            cell.border = _BORDER
            if c >= 4:
                cell.alignment = Alignment(horizontal="center")
        for col in (COL_OPENING, COL_RECEIVED, COL_RECEIVED_MCC, COL_TRANSFER, COL_MPP_SALE,
                    COL_DAMAGE, COL_EXPIRE, COL_CLOSING):
            totals[col] += values[col] or 0
        row_ptr += 1

    tcell = ws.cell(row_ptr, 1, "Total")
    tcell.font = Font(bold=True)
    for c in range(1, ncol + 1):
        cell = ws.cell(row_ptr, c)
        cell.border = _BORDER
        cell.fill = _TOTAL_FILL
        cell.font = Font(bold=True)
        col0 = c - 1
        if col0 in totals:
            cell.value = totals[col0]
            cell.alignment = Alignment(horizontal="center")

    widths = [7, 16, 34, 12, 18, 16, 16, 11, 13, 13, 13, 16]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A5"
    return ws


def _write_stock_block(ws, row, title_line, rows_data):
    """Write one report block (title line + column headers + data rows + Total) starting at
    `row`; return the next free row (after a blank gap). `rows_data` is a list of dicts with
    the report fields. Used for both per-cycle and month-summary blocks so the download looks
    exactly like the uploaded workbook (all cycles stacked in one sheet)."""
    ncol = len(REPORT_COLUMNS)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
    ws.cell(row, 1, title_line).font = Font(bold=True)
    row += 1
    for c, title in enumerate(REPORT_COLUMNS, start=1):
        cell = ws.cell(row, c, title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
    row += 1
    totals = defaultdict(int)
    for i, rd in enumerate(rows_data, start=1):
        vals = [i, rd["bmc"], rd["product"], rd["opening"], rd["received"], rd["received_mcc"],
                rd["transfer"], rd["sale"], rd["damage"], rd["expire"], rd["closing"], rd.get("remark", "")]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row, c, v)
            cell.border = _BORDER
            if c >= 4:
                cell.alignment = Alignment(horizontal="center")
        for col in (COL_OPENING, COL_RECEIVED, COL_RECEIVED_MCC, COL_TRANSFER, COL_MPP_SALE,
                    COL_DAMAGE, COL_EXPIRE, COL_CLOSING):
            totals[col] += vals[col] or 0
        row += 1
    ws.cell(row, 1, "Total").font = Font(bold=True)
    for c in range(1, ncol + 1):
        cell = ws.cell(row, c)
        cell.border = _BORDER
        cell.fill = _TOTAL_FILL
        cell.font = Font(bold=True)
        if (c - 1) in totals:
            cell.value = totals[c - 1]
            cell.alignment = Alignment(horizontal="center")
    return row + 3   # blank gap before the next block


def build_month_workbook(monthly_cycle, only_bmc=None, include_summary=True):
    """Build the downloadable workbook for a WHOLE month: one sheet per MCC/BMC with every
    cycle of the month stacked vertically (Cycle 1, 2, 3 ...) followed by a month-summary
    block — the same layout as the uploaded SMPCL file. Pass ``only_bmc`` for a single
    location; ``include_summary`` appends the company-wide 'SMPCL SUMMERY' sheet."""
    stmts = list(StockStatement.objects
                 .filter(cycle__monthly_cycle=monthly_cycle)
                 .select_related("cycle")
                 .order_by("cycle__cycle_number"))
    report_prods = report_products()
    report_pids = [p.id for p, _d in report_prods]
    month_display = f"{monthly_cycle.month}-{monthly_cycle.year}" if monthly_cycle else ""
    mon3 = monthly_cycle.month[:3] if monthly_cycle else ""

    # prefetch entries: (statement_id, bmc_id) -> {product_id: entry}
    ent = {}
    bmap = {}
    q = (StockStatementEntry.objects.filter(statement__in=stmts, product_id__in=report_pids)
         .select_related("bmc_or_mcc"))
    if only_bmc is not None:
        q = q.filter(bmc_or_mcc=only_bmc)
    for e in q:
        ent.setdefault((e.statement_id, e.bmc_or_mcc_id), {})[e.product_id] = e
        bmap[e.bmc_or_mcc_id] = e.bmc_or_mcc
    bmcs = [only_bmc] if only_bmc is not None else sorted(bmap.values(), key=lambda b: b.name)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used = set()
    if not bmcs or not stmts:
        ws = wb.create_sheet(_safe_sheet_title("No Data", used))
        ws["A1"] = "No statement data for this month."
        bio = io.BytesIO(); wb.save(bio); bio.seek(0)
        return bio

    def row_from_entry(bmc, e):
        return {"bmc": bmc.name, "product": None,
                "opening": e.opening_balance if e else 0, "received": e.received if e else 0,
                "received_mcc": e.received_mcc if e else 0, "transfer": e.stock_transfer if e else 0,
                "sale": e.mpp_sale if e else 0, "damage": e.damage if e else 0,
                "expire": e.expire if e else 0, "closing": e.closing_balance if e else 0,
                "remark": e.remark if e else ""}

    for bmc in bmcs:
        ws = wb.create_sheet(_safe_sheet_title(bmc.name, used))
        ncol = len(REPORT_COLUMNS)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
        ws.cell(1, 1, "Shwetdhara Milk Producer Company Limited").font = Font(bold=True, size=13)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
        ws.cell(2, 1, f"Input Sale & Stock Report  Month {month_display}").font = Font(bold=True)
        row = 3
        # one block per cycle
        for st in stmts:
            cyc = st.cycle
            rng = f"{cyc.start_date.strftime('%d')}-{cyc.end_date.strftime('%d')} {mon3}-{monthly_cycle.year}"
            emap = ent.get((st.id, bmc.id), {})
            rows_data = []
            for prod, display in report_prods:
                rd = row_from_entry(bmc, emap.get(prod.id))
                rd["product"] = display
                rows_data.append(rd)
            row = _write_stock_block(ws, row, f"Sale Report Cycle-{rng}  |  {bmc.name}", rows_data)
        # month summary block (opening from first cycle, closing from last, flows summed)
        first_map = ent.get((stmts[0].id, bmc.id), {})
        last_map = ent.get((stmts[-1].id, bmc.id), {})
        srows = []
        for prod, display in report_prods:
            o = first_map.get(prod.id)
            c = last_map.get(prod.id)
            rd = {"bmc": bmc.name, "product": display,
                  "opening": o.opening_balance if o else 0, "closing": c.closing_balance if c else 0,
                  "received": 0, "received_mcc": 0, "transfer": 0, "sale": 0, "damage": 0, "expire": 0, "remark": ""}
            for st in stmts:
                e = ent.get((st.id, bmc.id), {}).get(prod.id)
                if e:
                    rd["received"] += e.received; rd["received_mcc"] += e.received_mcc
                    rd["transfer"] += e.stock_transfer; rd["sale"] += e.mpp_sale
                    rd["damage"] += e.damage; rd["expire"] += e.expire
            srows.append(rd)
        _write_stock_block(ws, row, f"Month {month_display} Summary  |  {bmc.name}", srows)

        widths = [7, 22, 34, 12, 18, 16, 16, 11, 13, 13, 13, 16]
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A3"

    if include_summary and only_bmc is None:
        build_smpcl_summary(wb, stmts[0], used)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# ---------------------------------------------------------------------------------------
# Step 4: ingest the filled report (Damage / Expire) -> finalize + carry forward
# ---------------------------------------------------------------------------------------
def _to_int(v):
    try:
        return int(round(float(v))) if v not in (None, "") else 0
    except (TypeError, ValueError):
        return 0


def ingest_filled_report(statement, file_obj):
    """Read Damage/Expire back from the filled workbook, recompute Closing, finalize."""
    wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
    bmc_by_name = {b.name.strip().upper(): b for b in BMCOrMCC.objects.all()}
    prod_cache = build_product_norm_cache()

    updated = 0
    unmatched_rows = 0
    for ws in wb.worksheets:
        rows = _materialize(ws)   # read once; random cell access on read-only is O(n**2)
        # BMC/MCC for this sheet: prefer sheet title, else the MCC/BMC Name column.
        sheet_bmc = bmc_by_name.get(ws.title.strip().upper())
        # find header row (row containing 'Product Name')
        header_row = None
        for idx, row in enumerate(rows[:12]):
            if "product name" in [_norm_header(v) for v in row]:
                header_row = idx
                break
        if header_row is None:
            continue

        for r in range(header_row + 1, len(rows)):
            row = rows[r]
            name_cell = _cell(row, COL_PRODUCT + 1)
            first_cell = _cell(row, 1)
            if first_cell and str(first_cell).strip().lower() == "total":
                continue
            if not name_cell:
                continue
            bmc = sheet_bmc or bmc_by_name.get(str(_cell(row, 2) or "").strip().upper())
            if not bmc:
                unmatched_rows += 1
                continue
            product = resolve_product_by_name(name_cell, prod_cache, create=False)
            if not product:
                unmatched_rows += 1
                continue
            entry = statement.entries.filter(bmc_or_mcc=bmc, product=product).first()
            if not entry:
                unmatched_rows += 1
                continue
            entry.damage = _to_int(_cell(row, COL_DAMAGE + 1))
            entry.expire = _to_int(_cell(row, COL_EXPIRE + 1))
            remark = _cell(row, len(REPORT_COLUMNS))
            if remark:
                entry.remark = str(remark)[:255]
            entry.save()   # recomputes closing
            updated += 1

    statement.status = StockStatement.STATUS_FINALIZED
    statement.finalized_at = timezone.now()
    statement.save()
    rechain_month(statement.cycle.monthly_cycle)   # carry closings forward to later cycles
    return {"updated": updated, "unmatched_rows": unmatched_rows}


def ingest_full_report(file_obj):
    """Robust reader for a full Sale & Stock Report workbook.

    A location sheet stacks several blocks vertically — each introduced by a
    'Sale Report Cycle-DD-DD Mon-YYYY' line, then a column-header row, then data rows,
    ending at a 'Total' row. Any trailing block with NO cycle line (the month summary) is
    ignored. Columns are matched by header text, so single 'Received' / split
    'Received NDS + MCC' and combined 'Damage/Expire' / separate columns all work. BMC/MCC
    is matched loosely (sheet title or the MCC/BMC column). Product names resolve through the
    curated aliases, so 'Cattle Feed Sag (50_Kg)' updates 'CF- SAG 50 KG'.

    Fills every (cycle, location, product) it finds across the whole file, sets each touched
    statement to manual-override, and re-chains affected months. Returns a summary.
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
    bmc_by_norm = {_norm_bmc(b.name): b for b in BMCOrMCC.objects.all()}
    prod_cache = build_product_norm_cache()

    per_cycle = {}          # {cycle_id: {(bmc_id, product_id): field-values}}
    cyc_by_id = {}
    unmatched_locations = set()
    unmatched_rows = 0

    for ws in wb.worksheets:
        if ws.title.strip().upper().startswith("SMPCL"):
            continue   # summary sheet is derived, never a source
        sheet_bmc = bmc_by_norm.get(_norm_bmc(ws.title))
        rows = _materialize(ws)

        cur_cycle = None      # the cycle whose block we're inside
        cols = None           # column map for the current block
        in_data = False
        for row in rows:
            parsed = None
            for v in row[:3]:
                parsed = parse_cycle_header(v)
                if parsed:
                    break
            if parsed:         # 'Sale Report Cycle-…' line -> start of a cycle block
                sd, ed, month, yr = parsed
                cur_cycle = (Cycle.objects
                             .filter(monthly_cycle__month__iexact=month, monthly_cycle__year=yr,
                                     start_date__day=sd, end_date__day=ed)
                             .select_related("monthly_cycle").first())
                cols = None
                in_data = False
                continue
            hnorm = [re.sub(r"\s+", " ", str(v or "").strip().lower()) for v in row]
            if "product name" in hnorm:
                cols = _map_report_columns(row)
                in_data = True
                continue
            first = str(row[0] if row else "").strip().lower()
            if first == "total":
                cur_cycle = None   # block ended; a following block with no cycle line = summary -> skip
                cols = None
                in_data = False
                continue
            if not (in_data and cur_cycle and cols):
                continue

            def g(field):
                idx = cols.get(field)
                return row[idx - 1] if idx and idx - 1 < len(row) else None

            pname = g("product")
            if not pname or str(pname).strip().lower() in _PRODUCT_NAME_JUNK:
                continue
            bmc = sheet_bmc or bmc_by_norm.get(_norm_bmc(g("bmc")))
            if not bmc:
                unmatched_rows += 1
                unmatched_locations.add(str(g("bmc") or ws.title))
                continue
            product = resolve_product_by_name(pname, prod_cache, create=True)
            if not product:
                unmatched_rows += 1
                continue
            if "damage_combined" in cols:
                damage, expire = _to_int(g("damage_combined")), 0
            else:
                damage, expire = _to_int(g("damage")), _to_int(g("expire"))
            remark = g("remark")
            per_cycle.setdefault(cur_cycle.id, {})[(bmc.id, product.id)] = {
                "opening_balance": _to_int(g("opening")),
                "received": _to_int(g("received")),
                "received_mcc": _to_int(g("received_mcc")),
                "stock_transfer": _to_int(g("transfer")),
                "mpp_sale": _to_int(g("sale")),
                "damage": damage, "expire": expire,
                "remark": (str(remark)[:255] if remark else ""),
            }
            cyc_by_id[cur_cycle.id] = cur_cycle
    wb.close()

    # ---- apply, one statement per cycle ----
    updated = created = 0
    affected_months = {}
    cycles_filled = {}
    for cid, entries in per_cycle.items():
        cyc = cyc_by_id[cid]
        statement = StockStatement.objects.filter(cycle=cyc).first()
        if not statement:
            statement = generate_statement(cyc, user=None)   # seed the cycle's statement
        for (bmc_id, pid), vals in entries.items():
            entry = statement.entries.filter(bmc_or_mcc_id=bmc_id, product_id=pid).first()
            if entry is None:
                entry = StockStatementEntry(statement=statement, bmc_or_mcc_id=bmc_id, product_id=pid)
                created += 1
            for k, v in vals.items():
                setattr(entry, k, v)
            entry.save()   # recomputes closing
            updated += 1
        ensure_all_locations(statement)   # keep every location visible, even with no data
        statement.is_manual_override = True
        statement.override_uploaded_at = timezone.now()
        statement.save(update_fields=["is_manual_override", "override_uploaded_at", "updated_at"])
        cycles_filled[cyc.name] = len(entries)
        affected_months[cyc.monthly_cycle_id] = cyc.monthly_cycle

    for mc in affected_months.values():
        rechain_month(mc)

    return {
        "updated": updated, "created": created, "unmatched_rows": unmatched_rows,
        "cycles_filled": cycles_filled,
        "unmatched_locations": sorted(unmatched_locations),
    }
