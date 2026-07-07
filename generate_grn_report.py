"""
GRN Against PO Report - Oct 2025 to Mar 2026
All Locations | One sheet per location

Sections per location sheet:
  0 - Current Users at Location (active CustomUsers)
  0b- Historical GRN Performers (employees who did GRNs, incl. those who left)
  A - Requisition → PO → GRN full chain (the core new section)
  B - PO-wise Summary with best GRN
  C - Detailed GRN line items

Usage:
    pip install mysql-connector-python openpyxl
    python generate_grn_report.py
"""

import mysql.connector
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from collections import defaultdict

# ─── CONFIG ───────────────────────────────────────────────────────────────────
DB_CONFIG = dict(host='localhost', user='root', password='root@123', database='shwetdhara_db')
DATE_FROM = date(2025, 10, 1)
DATE_TO   = date(2026, 3, 31)

# ─── COLOURS ──────────────────────────────────────────────────────────────────
DARK_BLUE    = "1F3864"
MID_BLUE     = "2E75B6"
LIGHT_BLUE   = "BDD7EE"
ACCENT_TEAL  = "00B0F0"
PURPLE_HDR   = "4B2E83"
LAVENDER     = "E8E0F0"
DARK_GREEN   = "375623"
LIGHT_GREEN  = "E2EFDA"
AMBER_HDR    = "7F6000"
AMBER_LIGHT  = "FFF2CC"
WHITE        = "FFFFFF"
LIGHT_GREY   = "F2F2F2"
MED_GREY     = "D9D9D9"
DARK_TEXT    = "1A1A2E"
GREEN_OK     = "E2EFDA"
ORANGE_WARN  = "FFF2CC"
RED_LIGHT    = "FCE4D6"
RED_HDR      = "C00000"

LOCATIONS = [
    "BAHRAICH", "BARAUNSA", "AYODHYA H.O STORE", "NANPARA", "AKBARPUR",
    "RAEBARELI", "BADLAPUR", "RAM SANEHI GHAT", "RAJESULTANPUR", "COLONELGANJ",
    "PRATAPGARH", "AMETHI", "MAHRAJGANJ", "BALRAMPUR", "UMARWAL",
    "SUKULBAZAR", "MIHIPURWA", "MILKIPUR", "KURWAR",
]

# ─── STYLE HELPERS ────────────────────────────────────────────────────────────
def bd():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def mbd():
    m = Side(style='medium')
    return Border(left=m, right=m, top=m, bottom=m)

def hf(size=10, bold=True, color=WHITE):
    return Font(name='Arial', size=size, bold=bold, color=color)

def bf(size=10, bold=False, color=DARK_TEXT):
    return Font(name='Arial', size=size, bold=bold, color=color)

def fl(c):
    return PatternFill("solid", fgColor=c)

def cen():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def lft():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

def fmt_date(d):
    return d.strftime('%d-%b-%Y') if d else ''

def pct_str(rec, ord_):
    return f"{round(rec / ord_ * 100, 1)}%" if ord_ else '0.0%'

def section_header(ws, row, col_span, text, bg_color):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    c = ws.cell(row=row, column=1)
    c.value = text
    c.font = Font(name='Arial', size=11, bold=True, color=WHITE)
    c.fill = fl(bg_color)
    c.alignment = lft()
    ws.row_dimensions[row].height = 22

def table_header_row(ws, row, headers, bg_color=DARK_BLUE):
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=ci)
        c.value = h
        c.font = hf(10)
        c.fill = fl(bg_color)
        c.alignment = cen()
        c.border = bd()
    ws.row_dimensions[row].height = 28

def data_row(ws, row, vals, row_fill, left_cols=()):
    for ci, val in enumerate(vals, start=1):
        c = ws.cell(row=row, column=ci)
        c.value = val
        c.font = bf()
        c.fill = row_fill
        c.alignment = lft() if ci in left_cols else cen()
        c.border = bd()
    ws.row_dimensions[row].height = 18


# ─── DATA FETCHERS ────────────────────────────────────────────────────────────

def fetch_grn_data(conn):
    """GRN records in date range, keyed by location."""
    q = """
        SELECT
            g.id, g.grn_number, g.requisition_number, g.po_number,
            g.location, g.stock_item, g.uom,
            g.quantity_ordered, g.quantity_received, g.quantity_rejected,
            g.date_of_receipt, g.employee_name, g.employee_code,
            g.remarks, g.created_at
        FROM main_app_dogrn g
        WHERE g.date_of_receipt BETWEEN %s AND %s
        ORDER BY g.location, g.date_of_receipt, g.po_number
    """
    cur = conn.cursor(dictionary=True)
    cur.execute(q, (DATE_FROM, DATE_TO))
    rows = cur.fetchall()
    cur.close()
    data = {}
    for r in rows:
        loc = (r['location'] or 'UNKNOWN').strip().upper()
        data.setdefault(loc, []).append(r)
    return data


def fetch_requisitions_for_location(conn, location, req_numbers):
    """
    Fetch PurchaseRequisition records matching a location and a set of
    requisition numbers.  Also fetches HOD approval status via LEFT JOIN.
    """
    if not req_numbers:
        return {}
    placeholders = ','.join(['%s'] * len(req_numbers))
    q = f"""
        SELECT
            pr.requisition_number,
            pr.date_of_request,
            pr.employee_name     AS req_employee,
            pr.employee_code     AS req_emp_code,
            pr.department,
            pr.location          AS req_location,
            pr.stock_item        AS req_stock_item,
            pr.uom               AS req_uom,
            pr.quantity          AS req_quantity,
            pr.status            AS req_status,
            pr.po_number         AS req_po_number,
            pr.grn_done,
            pr.remark,
            pr.expected_delivery_date,
            ha.status            AS hod_approval_status,
            CONCAT(cu.first_name,' ',cu.last_name) AS hod_approver_name
        FROM main_app_purchaserequisition pr
        LEFT JOIN main_app_hodapproval ha
               ON ha.requisition_id = pr.id
        LEFT JOIN main_app_customuser cu
               ON cu.id = ha.approved_by_id
        WHERE pr.requisition_number IN ({placeholders})
        ORDER BY pr.requisition_number
    """
    cur = conn.cursor(dictionary=True)
    cur.execute(q, list(req_numbers))
    rows = cur.fetchall()
    cur.close()
    # key by requisition_number (take the first row if duplicates from join)
    result = {}
    for r in rows:
        rn = r['requisition_number']
        if rn not in result:
            result[rn] = r
    return result


def fetch_location_users(conn):
    """Active CustomUsers grouped by UPPER(location)."""
    q = """
        SELECT
            UPPER(TRIM(u.location))  AS location,
            TRIM(CONCAT(u.first_name,' ',u.last_name)) AS full_name,
            u.email, u.employee_code, u.department,
            u.is_hod, u.is_purchase, u.is_finance, u.is_logistic,
            u.is_active, u.is_staff
        FROM main_app_customuser u
        WHERE u.is_active = 1
        ORDER BY u.location, u.last_name, u.first_name
    """
    cur = conn.cursor(dictionary=True)
    cur.execute(q)
    rows = cur.fetchall()
    cur.close()
    out = {}
    for r in rows:
        loc = r['location'] or 'UNKNOWN'
        out.setdefault(loc, []).append(r)
    return out


def user_role(u):
    roles = []
    if u.get('is_hod'):      roles.append('HOD')
    if u.get('is_purchase'): roles.append('Purchase')
    if u.get('is_finance'):  roles.append('Finance')
    if u.get('is_logistic'): roles.append('Logistics')
    if u.get('is_staff'):    roles.append('Admin')
    return ', '.join(roles) if roles else 'Staff'


# ─── SHEET BUILDER ────────────────────────────────────────────────────────────

TOTAL_COLS = 14   # max columns used across all sections

def write_location_sheet(wb, loc_name, grn_rows, loc_users, req_map):
    ws = wb.create_sheet(title=loc_name[:31])
    COL = TOTAL_COLS

    # totals
    po_set  = set(r['po_number'] for r in grn_rows if r['po_number'])
    grn_set = set(r['grn_number'] for r in grn_rows if r['grn_number'])
    tot_ord = sum(r['quantity_ordered']  or 0 for r in grn_rows)
    tot_rec = sum(r['quantity_received'] or 0 for r in grn_rows)
    tot_rej = sum(r['quantity_rejected'] or 0 for r in grn_rows)
    fulf    = pct_str(tot_rec, tot_ord)

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells(f'A1:{get_column_letter(COL)}1')
    ws['A1'] = f"GRN AGAINST PO REPORT  —  {loc_name}"
    ws['A1'].font = Font(name='Arial', size=14, bold=True, color=WHITE)
    ws['A1'].fill = fl(DARK_BLUE); ws['A1'].alignment = cen()
    ws.row_dimensions[1].height = 32

    ws.merge_cells(f'A2:{get_column_letter(COL)}2')
    ws['A2'] = f"Period: October 2025 – March 2026  |  Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}"
    ws['A2'].font = Font(name='Arial', size=10, italic=True, color=WHITE)
    ws['A2'].fill = fl(MID_BLUE); ws['A2'].alignment = cen()
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8  # spacer

    # ── KPI strip ─────────────────────────────────────────────────────────────
    kpis = [('Total POs', len(po_set)), ('Total GRNs', len(grn_set)),
            ('Total Lines', len(grn_rows)), ('Qty Ordered', tot_ord),
            ('Qty Received', tot_rec), ('Qty Rejected', tot_rej),
            ('Fulfillment', fulf)]
    for i, (lbl, val) in enumerate(kpis):
        cs = 1 + i * 2
        ce = cs + 1
        ws.merge_cells(start_row=4, start_column=cs, end_row=4, end_column=ce)
        c = ws.cell(row=4, column=cs)
        c.value = lbl; c.font = hf(9); c.fill = fl(ACCENT_TEAL)
        c.alignment = cen(); c.border = bd()
        ws.merge_cells(start_row=5, start_column=cs, end_row=5, end_column=ce)
        v = ws.cell(row=5, column=cs)
        v.value = val; v.font = Font(name='Arial', size=11, bold=True, color=DARK_TEXT)
        v.fill = fl(LIGHT_BLUE); v.alignment = cen(); v.border = bd()
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[6].height = 8  # spacer

    cur_row = 7

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 0 — CURRENT LOCATION USERS
    # ══════════════════════════════════════════════════════════════════════════
    section_header(ws, cur_row, COL,
                   f'▶  SECTION 0 : Current Users at {loc_name}', PURPLE_HDR)
    cur_row += 1

    u_hdrs = ['#', 'Full Name', 'Email', 'Employee Code',
              'Department', 'Role / Access', 'Status']
    table_header_row(ws, cur_row, u_hdrs, PURPLE_HDR)
    cur_row += 1

    if loc_users:
        for ui, u in enumerate(loc_users, start=1):
            rfill = fl(LAVENDER) if ui % 2 == 0 else fl(WHITE)
            data_row(ws, cur_row, [
                ui,
                (u['full_name'] or '').strip(),
                u['email'] or '',
                u['employee_code'] or '',
                u['department'] or '',
                user_role(u),
                'Active',
            ], rfill, left_cols=(2, 3, 5))
            cur_row += 1
    else:
        ws.merge_cells(f'A{cur_row}:{get_column_letter(COL)}{cur_row}')
        c = ws.cell(row=cur_row, column=1)
        c.value = f'⚠  No active users currently assigned to location: {loc_name}'
        c.font = Font(name='Arial', size=10, italic=True, color=RED_HDR)
        c.fill = fl(RED_LIGHT); c.alignment = cen(); c.border = bd()
        cur_row += 1

    cur_row += 1  # spacer

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 0b — HISTORICAL GRN PERFORMERS (incl. left employees)
    # ══════════════════════════════════════════════════════════════════════════
    # Collect unique employees from GRN records for this location
    emp_grn_stats = defaultdict(lambda: {'grn_count': 0, 'qty_received': 0,
                                          'first_date': None, 'last_date': None,
                                          'emp_code': ''})
    for r in grn_rows:
        name = (r['employee_name'] or 'UNKNOWN').strip()
        s = emp_grn_stats[name]
        s['grn_count'] += 1
        s['qty_received'] += r['quantity_received'] or 0
        s['emp_code'] = r.get('employee_code') or s['emp_code']
        rd = r['date_of_receipt']
        if rd:
            if s['first_date'] is None or rd < s['first_date']: s['first_date'] = rd
            if s['last_date']  is None or rd > s['last_date']:  s['last_date']  = rd

    # Flag: is this person a current active user at this location?
    active_names = {(u['full_name'] or '').strip().upper() for u in loc_users}

    section_header(ws, cur_row, COL,
                   '▶  SECTION 0b : All Employees Who Performed GRNs at This Location (Oct 2025 – Mar 2026)',
                   DARK_GREEN)
    cur_row += 1

    h_hdrs = ['#', 'Employee Name', 'Employee Code', 'GRN Count',
              'Total Qty Received', 'First GRN Date', 'Last GRN Date', 'Current Status']
    table_header_row(ws, cur_row, h_hdrs, DARK_GREEN)
    cur_row += 1

    for hi, (name, s) in enumerate(
            sorted(emp_grn_stats.items(), key=lambda x: x[1]['last_date'] or date(1900,1,1), reverse=True),
            start=1):
        still_active = name.upper() in active_names
        rfill = fl(LIGHT_GREEN) if still_active else fl(ORANGE_WARN)
        status_label = '✅ Current Employee' if still_active else '⚠ Left / Transferred'
        data_row(ws, cur_row, [
            hi, name,
            s['emp_code'] or '—',
            s['grn_count'],
            s['qty_received'],
            fmt_date(s['first_date']),
            fmt_date(s['last_date']),
            status_label,
        ], rfill, left_cols=(2, 8))
        cur_row += 1

    cur_row += 1  # spacer

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION A — REQUISITION → PO → GRN FULL CHAIN
    # ══════════════════════════════════════════════════════════════════════════
    section_header(ws, cur_row, COL,
                   '▶  SECTION A : Requisition → PO → GRN Full Chain', AMBER_HDR)
    cur_row += 1

    a_hdrs = [
        'Req. No.', 'Req. Date', 'Raised By', 'Dept.',
        'Stock Item (Requisition)', 'Req. Qty', 'UOM',
        'PO Number', 'HOD Approval',
        'GRN Number', 'GRN Date', 'Qty Ordered', 'Qty Received', 'GRN Performed By'
    ]
    table_header_row(ws, cur_row, a_hdrs, AMBER_HDR)
    cur_row += 1

    # Group GRN rows by requisition_number
    req_to_grns = defaultdict(list)
    for r in grn_rows:
        rn = r['requisition_number'] or 'N/A'
        req_to_grns[rn].append(r)

    # Sort by requisition number
    for rn in sorted(req_to_grns.keys()):
        req_info = req_map.get(rn, {})
        grns_for_req = sorted(req_to_grns[rn],
                               key=lambda x: x['date_of_receipt'] or date(1900,1,1))

        req_date  = fmt_date(req_info.get('date_of_request'))
        raised_by = req_info.get('req_employee', '—') or '—'
        dept      = req_info.get('department', '—') or '—'
        req_item  = req_info.get('req_stock_item', '—') or '—'
        req_qty   = req_info.get('req_quantity', '—')
        req_uom   = req_info.get('req_uom', '—') or '—'
        po_num    = req_info.get('req_po_number', '—') or '—'
        hod_st    = req_info.get('hod_approval_status', '—') or '—'

        # Colour-code HOD approval
        if hod_st == 'APPROVED':
            hod_fill = fl(LIGHT_GREEN)
        elif hod_st in ('REJECTED',):
            hod_fill = fl(RED_LIGHT)
        else:
            hod_fill = fl(AMBER_LIGHT)

        for gi, g in enumerate(grns_for_req):
            rfill = fl(WHITE) if gi % 2 == 0 else fl(LIGHT_GREY)
            # Only show requisition info on first GRN row for this req
            if gi == 0:
                vals = [
                    rn, req_date, raised_by, dept,
                    req_item, req_qty, req_uom,
                    po_num, hod_st,
                    g['grn_number'] or '',
                    fmt_date(g['date_of_receipt']),
                    g['quantity_ordered']  or 0,
                    g['quantity_received'] or 0,
                    g['employee_name']     or '',
                ]
            else:
                # continuation rows — blank the requisition columns
                vals = [
                    '↑', '', '', '', '', '', '', '', '',
                    g['grn_number'] or '',
                    fmt_date(g['date_of_receipt']),
                    g['quantity_ordered']  or 0,
                    g['quantity_received'] or 0,
                    g['employee_name']     or '',
                ]

            for ci, val in enumerate(vals, start=1):
                c = ws.cell(row=cur_row, column=ci)
                c.value = val
                c.font  = bf(bold=(ci == 1 and gi == 0))
                # HOD column gets its own colour
                if ci == 9 and gi == 0:
                    c.fill = hod_fill
                else:
                    c.fill = rfill
                c.alignment = lft() if ci in (3, 4, 5, 14) else cen()
                c.border = bd()
            ws.row_dimensions[cur_row].height = 18
            cur_row += 1

    cur_row += 1  # spacer

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION B — PO-WISE SUMMARY
    # ══════════════════════════════════════════════════════════════════════════
    section_header(ws, cur_row, COL,
                   '▶  SECTION B : PO-Wise Summary  (Best GRN Performance)', MID_BLUE)
    cur_row += 1

    b_hdrs = ['PO Number', 'GRN Numbers', 'Products Covered',
              'Qty Ordered', 'Qty Received', 'Qty Rejected',
              'Fulfillment %', 'Best GRN', 'Last Receipt Date',
              'Performed By', 'Status']
    table_header_row(ws, cur_row, b_hdrs, MID_BLUE)
    cur_row += 1

    po_summary = defaultdict(lambda: {
        'grns': set(), 'total_ordered': 0, 'total_received': 0,
        'total_rejected': 0, 'products': set(), 'last_date': None,
        'employee': '', 'rows': []
    })
    for r in grn_rows:
        po = r['po_number'] or 'N/A'
        p  = po_summary[po]
        p['grns'].add(r['grn_number'] or '')
        p['total_ordered']  += r['quantity_ordered']  or 0
        p['total_received'] += r['quantity_received'] or 0
        p['total_rejected'] += r['quantity_rejected'] or 0
        p['products'].add(r['stock_item'] or '')
        rd = r['date_of_receipt']
        if rd and (p['last_date'] is None or rd > p['last_date']):
            p['last_date'] = rd
            p['employee'] = r['employee_name'] or ''
        p['rows'].append(r)

    for po_num, p in sorted(po_summary.items(),
                             key=lambda x: x[1]['last_date'] or date(1900,1,1),
                             reverse=True):
        tot = p['total_ordered']
        rec = p['total_received']
        pct = round(rec / tot * 100, 1) if tot else 0
        grn_counts = defaultdict(int)
        for r in p['rows']:
            grn_counts[r['grn_number'] or ''] += r['quantity_received'] or 0
        best = max(grn_counts, key=lambda g: grn_counts[g]) if grn_counts else ''
        status = 'COMPLETE' if pct >= 100 else ('PARTIAL' if pct > 0 else 'PENDING')
        rfill  = fl(GREEN_OK) if pct >= 100 else (fl(ORANGE_WARN) if pct > 0 else fl(RED_LIGHT))

        vals = [
            po_num,
            ', '.join(sorted(g for g in p['grns'] if g)),
            '\n'.join(sorted(p['products'])),
            tot, rec, p['total_rejected'],
            f"{pct}%", best,
            fmt_date(p['last_date']),
            p['employee'], status,
        ]
        for ci, val in enumerate(vals, start=1):
            c = ws.cell(row=cur_row, column=ci)
            c.value = val; c.font = bf()
            c.fill = rfill if ci > 1 else fl(LIGHT_GREY)
            c.alignment = lft() if ci in (2, 3, 10) else cen()
            c.border = bd()
        ws.row_dimensions[cur_row].height = max(18, 14 * len(p['products']))
        cur_row += 1

    cur_row += 1  # spacer

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION C — DETAILED GRN LINE ITEMS
    # ══════════════════════════════════════════════════════════════════════════
    section_header(ws, cur_row, COL,
                   '▶  SECTION C : Detailed GRN Line Items', DARK_BLUE)
    cur_row += 1

    c_hdrs = ['#', 'GRN Number', 'Requisition No.', 'PO Number',
              'Stock Item', 'UOM', 'Qty Ordered', 'Qty Received',
              'Qty Rejected', 'Completion %', 'Date of Receipt',
              'Performed By', 'Remarks']
    table_header_row(ws, cur_row, c_hdrs, DARK_BLUE)
    cur_row += 1

    for ri, r in enumerate(
            sorted(grn_rows, key=lambda x: (x['po_number'] or '', x['date_of_receipt'] or date(1900,1,1))),
            start=1):
        p = round((r['quantity_received'] or 0) / (r['quantity_ordered'] or 1) * 100, 1)
        rfill = fl(LIGHT_GREY) if ri % 2 == 0 else fl(WHITE)
        data_row(ws, cur_row, [
            ri,
            r['grn_number']         or '',
            r['requisition_number'] or '',
            r['po_number']          or '',
            r['stock_item']         or '',
            r['uom']                or '',
            r['quantity_ordered']   or 0,
            r['quantity_received']  or 0,
            r['quantity_rejected']  or 0,
            f"{p}%",
            fmt_date(r['date_of_receipt']),
            r['employee_name']      or '',
            r['remarks']            or '',
        ], rfill, left_cols=(5, 12, 13))
        cur_row += 1

    # ── Column widths ──────────────────────────────────────────────────────────
    widths = [4, 14, 14, 28, 34, 8, 12, 12, 12, 12, 16, 24, 30, 10]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = 'A8'


# ─── COVER SHEET ──────────────────────────────────────────────────────────────

def write_cover_sheet(wb, all_data, loc_users_map, report_date):
    ws = wb.active
    ws.title = 'Summary'
    COL = 12

    ws.merge_cells(f'A1:{get_column_letter(COL)}1')
    ws['A1'] = 'GRN AGAINST PO — ALL LOCATIONS REPORT'
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color=WHITE)
    ws['A1'].fill = fl(DARK_BLUE); ws['A1'].alignment = cen()
    ws.row_dimensions[1].height = 38

    ws.merge_cells(f'A2:{get_column_letter(COL)}2')
    ws['A2'] = f'Period: October 2025 – March 2026  |  Generated: {report_date}'
    ws['A2'].font = Font(name='Arial', size=10, italic=True, color=WHITE)
    ws['A2'].fill = fl(MID_BLUE); ws['A2'].alignment = cen()
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8

    hdrs = ['Location', 'Current Users', 'Historical GRN Employees',
            'Total POs', 'Unique GRNs', 'Total Lines',
            'Qty Ordered', 'Qty Received', 'Qty Rejected',
            'Fulfillment %', 'Status']
    for ci, h in enumerate(hdrs, start=1):
        c = ws.cell(row=4, column=ci)
        c.value = h; c.font = hf(); c.fill = fl(DARK_BLUE)
        c.alignment = cen(); c.border = bd()
    ws.row_dimensions[4].height = 26

    grand = dict(pos=set(), grns=set(), lines=0, ordered=0, received=0, rejected=0)

    for ri, loc in enumerate(LOCATIONS, start=5):
        rows      = all_data.get(loc, [])
        lu        = loc_users_map.get(loc, [])
        cur_names = '\n'.join(
            f"{u['full_name'].strip()} ({user_role(u)})" for u in lu
        ) or '—'
        hist_names = '\n'.join(sorted({
            (r['employee_name'] or '').strip() for r in rows
            if (r['employee_name'] or '').strip()
        })) or '—'

        if not rows:
            vals   = [loc, cur_names, hist_names, 0, 0, 0, 0, 0, 0, '—', 'NO DATA']
            rfill  = fl(LIGHT_GREY)
        else:
            po_set  = set(r['po_number'] for r in rows if r['po_number'])
            grn_set = set(r['grn_number'] for r in rows if r['grn_number'])
            tot_ord = sum(r['quantity_ordered']  or 0 for r in rows)
            tot_rec = sum(r['quantity_received'] or 0 for r in rows)
            tot_rej = sum(r['quantity_rejected'] or 0 for r in rows)
            pct     = round(tot_rec / tot_ord * 100, 1) if tot_ord else 0
            status  = 'COMPLETE' if pct >= 100 else ('PARTIAL' if pct > 0 else 'PENDING')
            rfill   = fl(GREEN_OK) if pct >= 100 else (fl(ORANGE_WARN) if pct > 0 else fl(RED_LIGHT))
            vals    = [loc, cur_names, hist_names, len(po_set), len(grn_set),
                       len(rows), tot_ord, tot_rec, tot_rej, f"{pct}%", status]
            grand['pos']      |= po_set;   grand['grns']     |= grn_set
            grand['lines']    += len(rows); grand['ordered']  += tot_ord
            grand['received'] += tot_rec;  grand['rejected']  += tot_rej

        num_lines = max(1, max(len(cur_names.split('\n')), len(hist_names.split('\n'))))
        for ci, val in enumerate(vals, start=1):
            c = ws.cell(row=ri, column=ci)
            c.value = val; c.font = bf(bold=(ci == 1))
            c.fill = rfill
            c.alignment = lft() if ci in (1, 2, 3) else cen()
            c.border = bd()
        ws.row_dimensions[ri].height = max(20, 15 * num_lines)

    gt = 5 + len(LOCATIONS) + 1
    ws.row_dimensions[gt].height = 26
    g_pct = round(grand['received'] / grand['ordered'] * 100, 1) if grand['ordered'] else 0
    for ci, val in enumerate(
        ['GRAND TOTAL', '', '', len(grand['pos']), len(grand['grns']),
         grand['lines'], grand['ordered'], grand['received'],
         grand['rejected'], f"{g_pct}%", ''], start=1):
        c = ws.cell(row=gt, column=ci)
        c.value = val
        c.font = Font(name='Arial', size=11, bold=True, color=WHITE)
        c.fill = fl(DARK_BLUE); c.alignment = cen(); c.border = bd()

    widths = [28, 34, 34, 10, 12, 12, 14, 14, 14, 12, 12]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A5'


# ─── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    print("Connecting to database …")
    conn = mysql.connector.connect(**DB_CONFIG)
    print("Connected.")

    print("Fetching GRN data (Oct 2025 – Mar 2026) …")
    all_data = fetch_grn_data(conn)
    print(f"  → {sum(len(v) for v in all_data.values())} GRN lines | {len(all_data)} locations")

    print("Fetching location users …")
    loc_users_map = fetch_location_users(conn)
    print(f"  → {sum(len(v) for v in loc_users_map.values())} active users")

    # Collect all requisition numbers across all locations
    all_req_numbers = set()
    for rows in all_data.values():
        for r in rows:
            if r['requisition_number']:
                all_req_numbers.add(r['requisition_number'])
    print(f"Fetching {len(all_req_numbers)} unique requisitions …")
    req_map = fetch_requisitions_for_location(conn, None, all_req_numbers)
    print(f"  → {len(req_map)} requisition records found")

    conn.close()

    wb = Workbook()
    report_date = datetime.now().strftime('%d-%b-%Y %H:%M')

    print("Building cover sheet …")
    write_cover_sheet(wb, all_data, loc_users_map, report_date)

    all_locs = list(LOCATIONS) + [l for l in all_data if l not in LOCATIONS]
    for i, loc in enumerate(all_locs, start=1):
        rows = all_data.get(loc, [])
        lu   = loc_users_map.get(loc, [])
        if not rows:
            print(f"  [{i:02d}] {loc:<26} — no GRN data, skipping")
            continue
        print(f"  [{i:02d}] {loc:<26} — {len(rows):>3} GRN lines | {len(lu)} current user(s)")
        write_location_sheet(wb, loc, rows, lu, req_map)

    out = f"GRN_PO_Report_Oct2025_Mar2026_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(out)
    print(f"\n✅  Saved → {out}")

if __name__ == '__main__':
    main()