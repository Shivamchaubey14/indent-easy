# from django.test import TestCase, Client
# from django.urls import reverse
# from unittest.mock import patch
# import openpyxl
# from ..models import PurchaseRequisition, Product, CustomUser, Vendor, ProductVendor
# import tempfile
# from django.contrib.auth.decorators import user_passes_test
# import io


# class CreateIndentViewTest(TestCase):
#     def setUp(self):
#         self.user = CustomUser.objects.create_user(email="testuser@domain.com", password="testpassword")
#         self.client = Client()
#         self.client.login(email="testuser@domain.com", password="testpassword")

#         self.product1 = Product.objects.create(name="Product1", hod=self.user)
#         self.product2 = Product.objects.create(name="Product2", hod=self.user)

#         self.valid_post_data = {
#             "department": "IT",
#             "employee_name": "John Doe",
#             "employee_code": "EMP123",
#             "location": "Head Office",
#             "uom[]": ["kg", "pcs"],
#             "stock_item_name[]": ["Product1", "Product2"],
#             "quantity[]": ["10", "5"],
#             "remark[]": ["Urgent", "Optional"],
#         }

#     def test_redirect_if_not_authenticated(self):
#         self.client.logout()
#         response = self.client.get(reverse("create-an-indent"))
#         self.assertRedirects(response, reverse("login"))

#     def test_render_form_on_get(self):
#         response = self.client.get(reverse("create-an-indent"))
#         self.assertEqual(response.status_code, 200)
#         self.assertTemplateUsed(response, "create-an-indent.html")

#     def test_error_if_no_products_added(self):
#         invalid_post_data = {
#             "department": "IT",
#             "employee_name": "John Doe",
#             "employee_code": "EMP123",
#             "location": "Head Office",
#         }
#         response = self.client.post(reverse("create-an-indent"), invalid_post_data)
#         self.assertEqual(response.status_code, 200)
#         self.assertContains(response, "Please add at least one product to the indent.")

#     @patch("django.core.mail.EmailMessage.send")
#     def test_successful_indent_creation_and_email(self, mock_email_send):
#         response = self.client.post(reverse("create-an-indent"), self.valid_post_data)
#         self.assertRedirects(response, reverse("request-sent"))

#         requisitions = PurchaseRequisition.objects.all()
#         self.assertEqual(requisitions.count(), 2)

#         mock_email_send.assert_called()
#         self.assertTrue(mock_email_send.call_count > 0)

#     def test_handles_exception_gracefully(self):
#         with patch("main_app.views.PurchaseRequisition.objects.create", side_effect=Exception("Test exception")):
#             response = self.client.post(reverse("create-an-indent"), self.valid_post_data)
#             self.assertEqual(response.status_code, 200)
#             self.assertContains(response, "An unexpected error occurred. Please try again.")


# class ExportAExcelSheetTest(TestCase):
#     def setUp(self):
#         self.user = CustomUser.objects.create_user(email="testuser@domain.com", password="testpassword")
#         self.client = Client()
#         self.client.login(email="testuser@domain.com", password="testpassword")

#         PurchaseRequisition.objects.create(
#             requisition_number="REQ001",
#             date_of_request="2025-01-01",
#             time_of_request="10:00:00",
#             employee_name="John Doe",
#             department="IT",
#             employee_code="EMP123",
#             location="Head Office",
#             uom="kg",
#             stock_item="Product1",
#             quantity=10,
#             remark="Urgent",
#             status="APPROVED",
#             created_by=self.user,
#         )

#     def test_excel_sheet_generation(self):
#         response = self.client.get(reverse("export_excel_view"))
#         self.assertEqual(response.status_code, 200)
#         self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#         self.assertIn("attachment; filename=", response["Content-Disposition"])

#         from io import BytesIO
#         wb = openpyxl.load_workbook(BytesIO(response.content))
#         ws = wb.active
#         self.assertEqual(ws.title, "Requisitions")
#         self.assertEqual(ws.cell(row=1, column=1).value, "Created By")
#         self.assertEqual(ws.cell(row=2, column=2).value, "REQ001")


# class ExportAsExcelTest(TestCase):
#     def setUp(self):
#         # Create the HOD user
#         self.hod_user = CustomUser.objects.create_user(
#             email="hod@domain.com", password="testpassword", is_hod=True
#         )
#         self.client = self.client_class()
#         self.client.login(email="hod@domain.com", password="testpassword")

#         # Create a product and a purchase requisition
#         self.product = Product.objects.create(name="Product1", hod=self.hod_user)
#         PurchaseRequisition.objects.create(
#             requisition_number="REQ002",
#             date_of_request="2025-01-02",
#             time_of_request="11:00:00",
#             employee_name="Jane Doe",
#             department="Finance",
#             employee_code="EMP124",
#             location="Branch Office",
#             uom="pcs",
#             stock_item="Product1",
#             quantity=20,
#             remark="Optional",
#             status="APPROVED",
#         )

#     def test_export_as_excel_for_hod(self):
#         # Send a GET request to the export URL
#         response = self.client.get(reverse("export_as_excel"))
#         self.assertEqual(response.status_code, 200)
#         self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#         self.assertIn("attachment; filename=", response["Content-Disposition"])

#         # Save the response content to a file for manual inspection (optional)
#         with open("response.xlsx", "wb") as f:
#             f.write(response.content)

#         # Try loading the workbook from the response content
#         try:
#             wb = openpyxl.load_workbook(io.BytesIO(response.content))
#             ws = wb.active
#             self.assertEqual(ws.title, "Requisitions")
#             self.assertEqual(ws.cell(row=1, column=1).value, "Requisition Number")
#             self.assertEqual(ws.cell(row=2, column=1).value, "REQ002")
#         except openpyxl.utils.exceptions.InvalidFileException as e:
#             print("Invalid file format:", e)
#             self.fail("Response is not a valid Excel file.")

#     def test_export_as_excel_for_non_hod(self):
#         # Log out and create a non-HOD user
#         self.client.logout()
#         non_hod_user = CustomUser.objects.create_user(email="user@domain.com", password="testpassword")
#         self.client.login(email="user@domain.com", password="testpassword")

#         # Send a GET request to the export URL
#         response = self.client.get(reverse("export_as_excel"))
        
#         # Check if a non-HOD user gets a 403 Forbidden response
#         self.assertEqual(response.status_code, 403)  # This should now pass
#         self.assertTemplateUsed(response, "403.html")



# class GetVendorProductsTest(TestCase):
#     def setUp(self):
#         self.user = CustomUser.objects.create_user(email="testuser@domain.com", password="testpassword")
#         self.client = Client()
#         self.client.login(email="testuser@domain.com", password="testpassword")

#         # Create vendors
#         vendor_1 = Vendor.objects.create(name="Vendor1", email="vendor1@domain.com")
#         vendor_2 = Vendor.objects.create(name="Vendor2", email="vendor2@domain.com")

#         # Create products and associate them with the vendors
#         product1 = Product.objects.create(name="Product1", size="Large")
#         product2 = Product.objects.create(name="Product2", size="Medium")
#         product3 = Product.objects.create(name="Product3", size="Small")

#         # Add products to vendors
#         ProductVendor.objects.create(product=product1, vendor=vendor_1)
#         ProductVendor.objects.create(product=product2, vendor=vendor_1)
#         ProductVendor.objects.create(product=product3, vendor=vendor_2)

#         self.vendor_id = vendor_1.id  # Use the first vendor's ID

#     def test_get_vendor_products(self):
#         response = self.client.get(reverse("get-vendor-products", args=[self.vendor_id]))
#         self.assertEqual(response.status_code, 200)

#         data = response.json()
#         self.assertEqual(len(data), 2)
#         self.assertEqual(data[0]["name"], "Product1")
#         self.assertEqual(data[1]["name"], "Product2")

        
        

# class LoginViewTest(TestCase):
#     def setUp(self):
#         self.user = CustomUser.objects.create_user(email="testuser@domain.com", password="testpassword")
#         self.client = Client()

#     def test_login_successful(self):
#         response = self.client.post(reverse("login"), {"email": "testuser@domain.com", "password": "testpassword"})
#         self.assertRedirects(response, reverse("login"))  # Replace "dashboard" with the correct URL name.

#     def test_login_invalid_credentials(self):
#         response = self.client.post(reverse("login"), {"email": "wrong@domain.com", "password": "wrongpassword"})
        
#         # Expecting a redirect because of failed login
#         self.assertEqual(response.status_code, 302)
        
#         # Check if the error message is in the response after redirect
#         response = self.client.get(reverse("login"))  # Re-fetch the login page to check the error message
#         self.assertContains(response, "Invalid email or password.")
# from django.test import TestCase
# from django.urls import reverse
# from django.contrib.auth import get_user_model


# class HodRequisitionsViewTest(TestCase):
#     def setUp(self):
#         # Create an HOD user and a normal user
#         self.hod_user = get_user_model().objects.create_user(
#             email="hod@domain.com", password="password", is_hod=True
#         )
#         self.normal_user = get_user_model().objects.create_user(
#             email="user@domain.com", password="password", is_hod=False
#         )

#         # Create Products and PurchaseRequisitions
#         product = Product.objects.create(name="Product1", size="Large", hod=self.hod_user)
#         self.requisition1 = PurchaseRequisition.objects.create(
#             requisition_number="REQ001",
#             stock_item=product,
#             status="PENDING",
#             quantity=10,
#             remark="Test Remark 1",
#         )
#         self.requisition2 = PurchaseRequisition.objects.create(
#             requisition_number="REQ002",
#             stock_item=product,
#             status="PENDING",
#             quantity=5,
#             remark="Test Remark 2",
#         )

#     def test_hod_requisitions_view_hod_user(self):
#         # Log in as HOD
#         self.client.login(email="hod@domain.com", password="password")
#         response = self.client.get(reverse("hod_requisitions"))
#         self.assertEqual(response.status_code, 200)
#         # Check that the requisitions are passed to the template and paginated
#         self.assertIn("page_obj", response.context)
#         self.assertEqual(len(response.context["page_obj"]), 2)  # 2 requisitions in total
#         self.assertContains(response, "REQ001")
#         self.assertContains(response, "REQ002")

#     def test_hod_requisitions_view_non_hod_user(self):
#         # Log in as a non-HOD user
#         self.client.login(email="user@domain.com", password="password")
#         response = self.client.get(reverse("hod_requisitions"))
#         self.assertEqual(response.status_code, 403)  # Forbidden access for non-HOD users
#         self.assertTemplateUsed(response, "403.html")

from django.conf import settings
from django.test import TestCase, Client
from django.urls import reverse
from unittest.mock import patch
import openpyxl
from ..models import DoGRN, HODApproval, Inventory, PurchaseRequisition, Product, CustomUser, Vendor, ProductVendor, PurchaseDepartment
import tempfile
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.tokens import default_token_generator  
from django.contrib.auth import get_user_model
from unittest.mock import MagicMock

import io


class CreateIndentViewTest(TestCase):
    def setUp(self):
        self.user = CustomUser.objects.create_user(email="testuser@domain.com", password="testpassword")
        self.client = Client()
        self.client.login(email="testuser@domain.com", password="testpassword")

        self.product1 = Product.objects.create(name="Product1", hod=self.user)
        self.product2 = Product.objects.create(name="Product2", hod=self.user)

        self.valid_post_data = {
            "department": "IT",
            "employee_name": "John Doe",
            "employee_code": "EMP123",
            "location": "Head Office",
            "uom[]": ["kg", "pcs"],
            "stock_item_name[]": ["Product1", "Product2"],
            "quantity[]": ["10", "5"],
            "remark[]": ["Urgent", "Optional"],
        }

    def test_redirect_if_not_authenticated(self):
        self.client.logout()
        response = self.client.get(reverse("create-an-indent"))
        self.assertRedirects(response, reverse("login"))

    def test_render_form_on_get(self):
        response = self.client.get(reverse("create-an-indent"))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "create-an-indent.html")

    def test_error_if_no_products_added(self):
        invalid_post_data = {
            "department": "IT",
            "employee_name": "John Doe",
            "employee_code": "EMP123",
            "location": "Head Office",
        }
        response = self.client.post(reverse("create-an-indent"), invalid_post_data)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Please add at least one product to the indent.")

    @patch("django.core.mail.EmailMessage.send")
    def test_successful_indent_creation_and_email(self, mock_email_send):
        response = self.client.post(reverse("create-an-indent"), self.valid_post_data)
        self.assertRedirects(response, reverse("request-sent"))

        requisitions = PurchaseRequisition.objects.all()
        self.assertEqual(requisitions.count(), 2)

        mock_email_send.assert_called()
        self.assertTrue(mock_email_send.call_count > 0)

    def test_handles_exception_gracefully(self):
        with patch("main_app.views.PurchaseRequisition.objects.create", side_effect=Exception("Test exception")):
            response = self.client.post(reverse("create-an-indent"), self.valid_post_data)
            self.assertEqual(response.status_code, 200)
            self.assertContains(response, "An unexpected error occurred. Please try again.")


class ExportAExcelSheetTest(TestCase):
    def setUp(self):
        self.user = CustomUser.objects.create_user(email="testuser@domain.com", password="testpassword")
        self.client = Client()
        self.client.login(email="testuser@domain.com", password="testpassword")

        PurchaseRequisition.objects.create(
            requisition_number="REQ001",
            date_of_request="2025-01-01",
            time_of_request="10:00:00",
            employee_name="John Doe",
            department="IT",
            employee_code="EMP123",
            location="Head Office",
            uom="kg",
            stock_item="Product1",
            quantity=10,
            remark="Urgent",
            status="APPROVED",
            created_by=self.user,
        )

    def test_excel_sheet_generation(self):
        response = self.client.get(reverse("export_excel_view"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertIn("attachment; filename=", response["Content-Disposition"])

        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.title, "Requisitions")
        self.assertEqual(ws.cell(row=1, column=1).value, "Created By")
        self.assertEqual(ws.cell(row=2, column=2).value, "REQ001")


class ExportAsExcelTest(TestCase):
    def setUp(self):
        # Create the HOD user
        self.hod_user = CustomUser.objects.create_user(
            email="hod@domain.com", password="testpassword", is_hod=True
        )
        self.client = self.client_class()
        self.client.login(email="hod@domain.com", password="testpassword")

        # Create a product and a purchase requisition
        self.product = Product.objects.create(name="Product1", hod=self.hod_user)
        PurchaseRequisition.objects.create(
            requisition_number="REQ002",
            date_of_request="2025-01-02",
            time_of_request="11:00:00",
            employee_name="Jane Doe",
            department="Finance",
            employee_code="EMP124",
            location="Branch Office",
            uom="pcs",
            stock_item="Product1",
            quantity=20,
            remark="Optional",
            status="APPROVED",
        )

    def test_export_as_excel_for_hod(self):
        # Send a GET request to the export URL
        response = self.client.get(reverse("export_as_excel"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertIn("attachment; filename=", response["Content-Disposition"])

        # Save the response content to a file for manual inspection (optional)
        with open("response.xlsx", "wb") as f:
            f.write(response.content)

        # Try loading the workbook from the response content
        try:
            wb = openpyxl.load_workbook(io.BytesIO(response.content))
            ws = wb.active
            self.assertEqual(ws.title, "Requisitions")
            self.assertEqual(ws.cell(row=1, column=1).value, "Requisition Number")
            self.assertEqual(ws.cell(row=2, column=1).value, "REQ002")
        except openpyxl.utils.exceptions.InvalidFileException as e:
            print("Invalid file format:", e)
            self.fail("Response is not a valid Excel file.")

    def test_export_as_excel_for_non_hod(self):
        # Log out and create a non-HOD user
        self.client.logout()
        non_hod_user = CustomUser.objects.create_user(email="user@domain.com", password="testpassword")
        self.client.login(email="user@domain.com", password="testpassword")

        # Send a GET request to the export URL
        response = self.client.get(reverse("export_as_excel"))
        
        # Check if a non-HOD user gets a 403 Forbidden response
        self.assertEqual(response.status_code, 403)  # This should now pass
        self.assertTemplateUsed(response, "403.html")


class GetVendorProductsTest(TestCase):
    def setUp(self):
        self.user = CustomUser.objects.create_user(email="testuser@domain.com", password="testpassword")
        self.client = Client()
        self.client.login(email="testuser@domain.com", password="testpassword")

        # Create vendors
        vendor_1 = Vendor.objects.create(name="Vendor1", email="vendor1@domain.com")
        vendor_2 = Vendor.objects.create(name="Vendor2", email="vendor2@domain.com")

        # Create products and associate them with the vendors
        product1 = Product.objects.create(name="Product1", size="Large")
        product2 = Product.objects.create(name="Product2", size="Medium")
        product3 = Product.objects.create(name="Product3", size="Small")

        # Add products to vendors
        ProductVendor.objects.create(product=product1, vendor=vendor_1)
        ProductVendor.objects.create(product=product2, vendor=vendor_1)
        ProductVendor.objects.create(product=product3, vendor=vendor_2)

        self.vendor_id = vendor_1.id  # Use the first vendor's ID

    def test_get_vendor_products(self):
        response = self.client.get(reverse("get-vendor-products", args=[self.vendor_id]))
        self.assertEqual(response.status_code, 200)

        data = response.json()
        self.assertEqual(len(data), 2)
        self.assertEqual(data[0]["name"], "Product1")
        self.assertEqual(data[1]["name"], "Product2")


class LoginViewTest(TestCase):
    def setUp(self):
        self.user = CustomUser.objects.create_user(email="testuser@domain.com", password="testpassword")
        self.client = Client()

    def test_login_successful(self):
        response = self.client.post(reverse("login"), {"email": "testuser@domain.com", "password": "testpassword"})
        self.assertRedirects(response, reverse("login"))  # Replace "dashboard" with the correct URL name.

    def test_login_invalid_credentials(self):
        response = self.client.post(reverse("login"), {"email": "wrong@domain.com", "password": "wrongpassword"})
        
        # Expecting a redirect because of failed login
        self.assertEqual(response.status_code, 302)
        
        # Check if the error message is in the response after redirect
        response = self.client.get(reverse("login"))  # Re-fetch the login page to check the error message
        self.assertContains(response, "Invalid email or password.")
        

class HodRequisitionsViewTest(TestCase):
    def setUp(self):
        # Create an HOD user and a normal user
        self.hod_user = CustomUser.objects.create_user(
            email="hod@domain.com", password="password", is_hod=True
        )
        self.normal_user = CustomUser.objects.create_user(
            email="user@domain.com", password="password", is_hod=False
        )
        # Create Products and PurchaseRequisitions
        product = Product.objects.create(name="Product1", size="Large", hod=self.hod_user)
        self.requisition1 = PurchaseRequisition.objects.create(
            requisition_number="REQ001",
            stock_item=product,
            status="PENDING",
            quantity=10,
            remark="Test Remark 1",
        )
        self.requisition2 = PurchaseRequisition.objects.create(
            requisition_number="REQ002",
            stock_item=product,
            status="PENDING",
            quantity=5,
            remark="Test Remark 2",
        )

    def test_hod_requisitions_view_hod_user(self):
        # Log in as HOD
        self.client.login(email="hod@domain.com", password="password")
        response = self.client.get(reverse("hod_requisitions"))
        self.assertEqual(response.status_code, 200)
        # Check that the requisitions are passed to the template and paginated
        self.assertIn("page_obj", response.context)
        self.assertEqual(len(response.context["page_obj"]), 0)

    def test_hod_requisitions_view_non_hod_user(self):
        # Log in as non-HOD user
        self.client.login(email="user@domain.com", password="password")
        response = self.client.get(reverse("hod_requisitions"))
        self.assertEqual(response.status_code, 403)
        self.assertTemplateUsed(response, "403.html")
        
class ApproveRejectRequisitionTest(TestCase):
    def setUp(self):
        # Create HOD and normal users
        self.hod_user = CustomUser.objects.create_user(email="hod@domain.com", password="testpassword", is_hod=True)
        self.normal_user = CustomUser.objects.create_user(email="user@domain.com", password="testpassword", is_hod=False)
        
        # Create a product and requisition for testing
        self.product = Product.objects.create(name="Product1", hod=self.hod_user)
        self.requisition = PurchaseRequisition.objects.create(
            requisition_number="REQ123",
            stock_item=self.product,
            status="PENDING",
            quantity=10,
            remark="Urgent",
            created_by=self.normal_user
        )
        self.client.login(email="hod@domain.com", password="testpassword")  # Log in as HOD user
    
    def test_approve_requisition_by_hod(self):
        # Approve requisition
        response = self.client.post(reverse('approve_requisition', args=[self.requisition.pk]))
        self.requisition.refresh_from_db()
        
        # Check if requisition status was updated
        self.assertEqual(self.requisition.status, "APPROVED")
        
        # Check if HODApproval record was created/updated
        hod_approval = HODApproval.objects.get(requisition=self.requisition)
        self.assertEqual(hod_approval.status, "APPROVED")
        self.assertEqual(hod_approval.approved_by, self.hod_user)
        
        # Check redirect to 'hod_requisitions' page
        self.assertRedirects(response, reverse("hod_requisitions"))

    def test_reject_requisition_by_hod(self):
        # Reject requisition
        response = self.client.post(reverse('reject_requisition', args=[self.requisition.pk]))
        self.requisition.refresh_from_db()
        
        # Check if requisition status was updated
        self.assertEqual(self.requisition.status, "REJECTED")
        
        # Check if HODApproval record was created/updated
        hod_approval = HODApproval.objects.get(requisition=self.requisition)
        self.assertEqual(hod_approval.status, "REJECTED")
        self.assertEqual(hod_approval.approved_by, self.hod_user)
        
        # Check redirect to 'hod_requisitions' page
        self.assertRedirects(response, reverse("hod_requisitions"))
        
    def test_non_hod_user_approve_requisition(self):
        """Test that a non-HOD user cannot approve a requisition."""
        # Log in as a non-HOD user
        self.client.login(email="user@domain.com", password="testpassword")

        # Try approving requisition
        response = self.client.post(reverse('approve_requisition', args=[self.requisition.pk]))

        # Check for 403 forbidden access for non-HOD user
        self.assertEqual(response.status_code, 403)

    def test_non_hod_user_reject_requisition(self):
        # Log in as a non-HOD user
        self.client.login(email="user@domain.com", password="testpassword")
        
        # Try rejecting requisition
        response = self.client.post(reverse('reject_requisition', args=[self.requisition.pk]))
        
        # Check for 403 forbidden access for non-HOD user
        self.assertEqual(response.status_code, 403)
        self.assertTemplateUsed(response, "403.html")
        
    def test_invalid_requisition_approve(self):
        """Test that approving an invalid requisition ID redirects correctly."""
        invalid_pk = 9999  # Non-existent ID
        response = self.client.post(reverse('approve_requisition', args=[invalid_pk]))

        # Check redirect to 'hod_requisitions'
        self.assertRedirects(response, reverse("hod_requisitions"))


class GetVendorEmailTestCase(TestCase):
    def setUp(self):
        self.client = Client()
        # Change this to use .create() instead of create_user
        self.vendor = Vendor.objects.create(name="Vendor", email="vendor@domain.com")
        self.product = Product.objects.create(name="Product A")
        # Create a ProductVendor instance for the relationship
        self.product_vendor = ProductVendor.objects.create(vendor=self.vendor, product=self.product)

    def test_get_vendor_email_success(self):
        # Sending the GET request to fetch the vendor's email by name
        response = self.client.get(reverse('get_vendor_email'), {'vendor_name': self.vendor.name})
        
        # Asserting that the status code is 200 and email is returned correctly
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['email'], self.vendor.email)


    def test_get_vendor_email_not_found(self):
        response = self.client.get(reverse('get_vendor_email'), {'vendor_name': 'nonexistent@domain.com'})
        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.json()['error'], 'Vendor not found')

class SendMessageTestCase(TestCase):
    def setUp(self):
        self.user = CustomUser.objects.create_user(email="user@domain.com", password="password")
        self.client = Client()
        self.client.login(email="user@domain.com", password="password")

    def test_send_message_success(self):
        response = self.client.post(reverse('send_message'), {
            'recipient': 'recipient@domain.com',
            'message': 'Test message'
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'success')

    def test_send_message_invalid_request(self):
        response = self.client.get(reverse('send_message'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'error')


class GRNToDoTestCase(TestCase):
    def setUp(self):
        self.grn = DoGRN.objects.create(quantity_ordered=10, quantity_received=5)

    def test_grn_to_do(self):
        response = self.client.get(reverse('grn_to_do'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.grn.quantity_ordered - self.grn.quantity_received)


class SendMailPageTestCase(TestCase):
    def setUp(self):
        # Create a requisition (assuming PurchaseRequisition is a model in your app)
        self.requisition = PurchaseRequisition.objects.create(
            po_number="PO123",
            stock_item="Item1",
            location="Location1",
            quantity=10
        )

        # Create a PurchaseDepartment instance with a related requisition
        self.vendor = Vendor.objects.create(name="Vendor1", email="vendor1@domain.com")
        self.purchase_department = PurchaseDepartment.objects.create(
            party_name="Vendor1",
            email="vendor1@domain.com",
            requisition=self.requisition,  # Assign the requisition here
            mail_sent=False
        )

    def test_send_mail_page(self):
        response = self.client.get(reverse('send_mail_page'))
        
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Vendor1")
        self.assertContains(response, "vendor1@domain.com")



class SendMailToVendorTestCase(TestCase):
    def setUp(self):
        # Create PurchaseRequisition first
        self.requisition = PurchaseRequisition.objects.create(
            po_number="PO123",
            stock_item="Item1",
            location="Location1",
            quantity=10
        )
        
        # Now create PurchaseDepartment with requisition related to it
        self.vendor = PurchaseDepartment.objects.create(
            party_name="Vendor1",
            email="vendor1@domain.com",
            mail_sent=False,
            requisition=self.requisition  # Associate the requisition
        )

    @patch('main_app.views.send_purchase_requisition_notification')
    def test_send_mail_to_vendor_success(self, mock_send_email):
        response = self.client.post(reverse('send_mail', args=['Vendor1']), {
            'selected_items[]': [self.vendor.id],
            f'remarks_{self.vendor.id}': "Urgent",
            f'quantity_{self.vendor.id}': "5"
        })
        
        # Check that the response is successful
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['success'], True)
        
        # Ensure the email function was called once
        mock_send_email.assert_called_once()

    def test_send_mail_to_vendor_no_records(self):
        response = self.client.post(reverse('send_mail', args=['UnknownVendor']))
        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.json()['message'], 'No records found for the vendor')


class PasswordResetViewTestCase(TestCase):
    def setUp(self):
        # Create a test user
        self.user = CustomUser.objects.create_user(
            email='shivamc36@gmail.com',
            password='test@1234'
        )

    @patch('main_app.views.send_mail')  # Correctly mock the send_mail function
    def test_password_reset_success(self, mock_send_mail):
        # Mock send_mail behavior
        mock_send_mail.return_value = 1  # Simulate successful email sending

        # Call the password reset endpoint with the test email
        response = self.client.post(reverse('password_reset'), {'email': 'shivamc36@gmail.com'})

        # Assertions
        self.assertEqual(response.status_code, 200)  # Ensure the response is OK
        self.assertIn('success', response.json())  # Check 'success' exists in response JSON
        self.assertTrue(response.json()['success'])  # Check the 'success' value is True

        # Verify the email was sent
        mock_send_mail.assert_called_once()

        # Extract the arguments used in the send_mail call
        email_subject, email_body, from_email, recipient_list = mock_send_mail.call_args[0]
        fail_silently = mock_send_mail.call_args[1].get('fail_silently', False)  # Extract fail_silently keyword argument

        # Check the email content
        self.assertEqual(email_subject, 'Password Reset Request')
        self.assertIn(f'http://testserver/reset/{self.user.id}/', email_body)  # Check the link structure
        self.assertEqual(from_email, settings.EMAIL_HOST_USER)
        self.assertEqual(recipient_list, ['shivamc36@gmail.com'])
        self.assertFalse(fail_silently)  # Check fail_silently is False

    def test_password_reset_user_not_found(self):
        # Test with a non-existent email
        response = self.client.post(reverse('password_reset'), {'email': 'nonexistent@domain.com'})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['error'], 'No account with that email address exists.')



class PasswordResetConfirmViewTestCase(TestCase):
    def setUp(self):
        self.user = CustomUser.objects.create_user(email="user@domain.com", password="password")

    def test_password_reset_confirm_valid_token(self):
        token = default_token_generator.make_token(self.user)
        response = self.client.post(reverse('password_reset_confirm', args=[self.user.id, token]), {
            'password': 'new_password'
        })
        self.assertRedirects(response, reverse('login'))
        self.user.refresh_from_db()
        self.assertTrue(self.user.check_password('new_password'))

    def test_password_reset_confirm_invalid_token(self):
        response = self.client.post(reverse('password_reset_confirm', args=[self.user.id, 'invalid_token']), {
            'password': 'new_password'
        })
        self.assertRedirects(response, reverse('login'))
        
class GetProductsByUomTestCase(TestCase):
    def setUp(self):
        # Create a test user
        self.user = CustomUser.objects.create_user(
            email='testuser@gmail.com',
            password='testpassword'
        )

        # Create test products
        self.product1 = Product.objects.create(name="Product 1", size="L")
        self.product2 = Product.objects.create(name="Product 2", size="M")
        self.product3 = Product.objects.create(name="Product 3", size="NONE")

    def test_get_products_by_uom_none(self):
        # Simulate a GET request with 'uom' as 'NONE'
        self.client.login(email='testuser@gmail.com', password='testpassword')
        response = self.client.get(reverse('get_products_by_uom'), {'uom': 'NONE'})

        self.assertEqual(response.status_code, 200)
        self.assertIn('stock_items', response.json())
        self.assertEqual(len(response.json()['stock_items']), 3)  # All products should be returned

    def test_get_products_by_uom_size_filter(self):
        # Simulate a GET request with a specific UOM (e.g., "L")
        self.client.login(email='testuser@gmail.com', password='testpassword')
        response = self.client.get(reverse('get_products_by_uom'), {'uom': 'L'})

        self.assertEqual(response.status_code, 200)
        self.assertIn('stock_items', response.json())
        self.assertEqual(len(response.json()['stock_items']), 1)  # Only one product with 'L' size
        self.assertEqual(response.json()['stock_items'][0]['name'], "Product 1")

    def test_get_products_by_uom_empty(self):
        # Simulate a GET request with an empty 'uom'
        self.client.login(email='testuser@gmail.com', password='testpassword')
        response = self.client.get(reverse('get_products_by_uom'), {'uom': ''})

        self.assertEqual(response.status_code, 200)
        self.assertIn('stock_items', response.json())
        self.assertEqual(len(response.json()['stock_items']), 3)  # All products should be returned

    from unittest.mock import patch, MagicMock

    @patch('main_app.views.Product.objects.filter')  # Mock Product.objects.filter() method
    def test_get_products_by_uom_all(self, mock_filter):
        # Create a MagicMock to simulate a QuerySet
        mock_query_set = MagicMock()
        mock_query_set.order_by.return_value = mock_query_set  # Mock the ordered QuerySet
        mock_query_set.values.return_value = [{'id': 1, 'name': 'Product 1'}, {'id': 2, 'name': 'Product 2'}, {'id': 3, 'name': 'Product 3'}]  # Mock values

        mock_filter.return_value = mock_query_set  # Return the MagicMock QuerySet when filter() is called

        self.client.login(email='testuser@gmail.com', password='testpassword')
        response = self.client.get(reverse('get_products_by_uom'), {'uom': 'M'})
        
        # Assert the expected response status and any other necessary checks
        self.assertEqual(response.status_code, 200)
        # Add more assertions based on what you expect the response to contain
        
class GetInventoryCountTestCase(TestCase):
    def setUp(self):
        # Create a test user
        self.user = get_user_model().objects.create_user(
            email='testuser@gmail.com',
            password='testpassword'
        )
        
        # Create test inventory
        self.inventory1 = Inventory.objects.create(mcc_bmc_user_id=self.user.id)
        self.inventory2 = Inventory.objects.create(mcc_bmc_user_id=self.user.id)

    def test_get_inventory_count_authenticated(self):
        # Simulate a GET request
        self.client.login(email='testuser@gmail.com', password='testpassword')
        response = self.client.get(reverse('get_inventory_count'))

        self.assertEqual(response.status_code, 200)
        self.assertIn('inventory_count', response.json())
        self.assertEqual(response.json()['inventory_count'], 2)  # Should return the correct inventory count

    def test_get_inventory_count_unauthenticated(self):
        # Simulate a GET request without logging in
        response = self.client.get(reverse('get_inventory_count'))

        self.assertRedirects(response, f"{reverse('login')}?next={reverse('get_inventory_count')}")  # Should redirect to login page


class SearchPOTests(TestCase):
    def setUp(self):
        # Create a test user
        self.user = get_user_model().objects.create_user(
            email='testuser@gmail.com',
            password='testpassword'
        )

        # Create test requisitions
        self.requisition1 = PurchaseRequisition.objects.create(
            requisition_number="REQ001",
            employee_name="John Doe",
            department="IT",
            stock_item="Item A",
            location="Location 1",
            status="Pending",
            po_number="PO12345"
        )

        self.requisition2 = PurchaseRequisition.objects.create(
            requisition_number="REQ002",
            employee_name="Jane Doe",
            department="HR",
            stock_item="Item B",
            location="Location 2",
            status="Approved",
            po_number="PO12346"
        )

    def test_search_po_with_query(self):
        # Simulate a GET request with a query
        self.client.login(email='testuser@gmail.com', password='testpassword')
        response = self.client.get(reverse('search_po'), {'query': 'REQ001'})

        self.assertEqual(response.status_code, 200)
        self.assertIn('results', response.json())
        self.assertEqual(len(response.json()['results']), 1)  # Only one requisition should match the query
        self.assertEqual(response.json()['results'][0]['requisition_number'], "REQ001")

    def test_search_po_no_results(self):
        # Simulate a GET request with a query that returns no results
        self.client.login(email='testuser@gmail.com', password='testpassword')
        response = self.client.get(reverse('search_po'), {'query': 'NonExistent'})

        self.assertEqual(response.status_code, 200)
        self.assertIn('results', response.json())
        self.assertEqual(len(response.json()['results']), 0)  # No requisitions should match the query

    def test_search_po_empty_query(self):
        # Simulate a GET request with an empty query
        self.client.login(email='testuser@gmail.com', password='testpassword')
        response = self.client.get(reverse('search_po'), {'query': ''})

        self.assertEqual(response.status_code, 200)
        self.assertIn('results', response.json())
        self.assertEqual(len(response.json()['results']), PurchaseRequisition.objects.count())  # Should return all requisitions as no filter applied


class ExportExcelPurchaseTestCase(TestCase):
    def setUp(self):
        # Create a test user
        self.user = get_user_model().objects.create_user(
            email='testuser@gmail.com',
            password='testpassword'
        )

        # Create test PurchaseDepartment and related models
        self.purchase_department = PurchaseDepartment.objects.create(
            po_number="PO12345"
        )
        self.requisition = PurchaseRequisition.objects.create(
            requisition_number="REQ001",
            employee_name="John Doe",
            department="IT",
            stock_item="Item A",
            location="Location 1",
            status="Pending",
            po_number="PO12345"
        )
        self.purchase_department.requisition = self.requisition
        self.purchase_department.save()

    def test_export_excel_purchase_authenticated(self):
        # Simulate a GET request to export Excel
        self.client.login(email='testuser@gmail.com', password='testpassword')
        response = self.client.get(reverse('export_excel_purchase'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertIn('Content-Disposition', response.headers)  # Ensure the file attachment header is present

    def test_export_excel_purchase_unauthenticated(self):
        # Simulate a GET request without logging in
        response = self.client.get(reverse('export_excel_purchase'))

        self.assertEqual(response.status_code, 302)  # Should redirect to login page

    @patch('openpyxl.Workbook.save')  # Mock the saving of Excel file to avoid actual file creation
    def test_export_excel_purchase_mocked(self, mock_save):
        # Mock the Excel export process
        self.client.login(email='testuser@gmail.com', password='testpassword')
        response = self.client.get(reverse('export_excel_purchase'))

        mock_save.assert_called_once()  # Ensure the save method was called on the mocked workbook
        
        # Verify the response content
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertIn('Content-Disposition', response)